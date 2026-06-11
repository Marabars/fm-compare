"""
KPI extractor: finds KPI values in workbook sheets using dictionary rules.
Searches by row label patterns, not just coordinates.
"""
from __future__ import annotations
import re
from typing import Any

from fm_compare.core.excel_reader import WorkbookData, SheetData
from fm_compare.core.business_dictionary import BusinessDictionary, KPIEntry
from fm_compare.core.models import CellAddress, KPIValue, KPIResolution
from fm_compare.core.utils import is_numeric
from fm_compare.security import safe_logger as log


_UNIT_RE = re.compile(
    r"\b(тыс\.?\s*руб|млн\.?\s*руб|млрд\.?\s*руб|тыс|млн|млрд|руб\.?|%|шт\.?|чел\.?|кг|т\b|м2|м²|ед\.?)\b",
    re.IGNORECASE | re.UNICODE,
)


def _find_unit_in_row(sd: SheetData, row: int, label_col: int) -> str:
    """
    Scan the KPI row for a unit-of-measure string.
    Checks columns 1 through label_col+1 (the cell just right of the label
    is included because some layouts put the unit there).
    Returns the shortest matching cell value; empty string if nothing found.
    """
    candidates: list[tuple[int, str]] = []
    for col in range(1, label_col + 2):
        cd = sd.cells.get((row, col))
        if cd is None or cd.value is None:
            continue
        s = str(cd.value).strip()
        if s and _UNIT_RE.search(s):
            candidates.append((len(s), s))
    if candidates:
        candidates.sort()
        return candidates[0][1]
    return ""


def _find_summary_col(sd: SheetData) -> int | None:
    """Find the column that likely contains summary/total values."""
    for col in range(3, min(sd.max_col + 1, 20)):
        non_empty = sum(1 for row in range(1, min(sd.max_row + 1, 50))
                        if sd.cells.get((row, col)) and
                        is_numeric(sd.cells[(row, col)].value))
        if non_empty > 3:
            return col
    return 3


def _find_kpi_row(
    sd: SheetData,
    kpi: KPIEntry,
    label_col: int = 2,
) -> tuple[int, Any] | None:
    """
    Find best-matching row for a KPI search pattern.
    Scores candidates by match coverage (matched chars / label length): a
    label that consists mostly of the matched fragment scores higher than a
    long label that merely contains it.  Ties are broken by row position
    (earlier wins) so that summary rows near the top of a sheet are preferred.
    Returns (row_index, label_cell_value) or None.
    """
    try:
        pattern = re.compile(kpi.search_pattern, re.IGNORECASE | re.UNICODE)
    except re.error:
        return None

    best_row: int | None = None
    best_label: Any = None
    best_score: float = -1.0

    for row in range(1, sd.max_row + 1):
        cd = sd.cells.get((row, label_col))
        if cd is None or cd.value is None:
            continue
        label = str(cd.value).strip()
        if not label:
            continue
        m = pattern.search(label)
        if not m:
            continue
        # Score = start_bonus + coverage.
        # start_bonus = 1.0 when the match begins at position 0 (the label
        # _starts_ with the KPI term, e.g. "Выручка от реализации").  This
        # guarantees any start-of-label match beats any mid-label match, even
        # when mid-label coverage is numerically higher (e.g. "Период продаж").
        # Coverage = matched_chars / label_len rewards concise, precise labels.
        coverage = len(m.group()) / len(label)
        start_bonus = 1.0 if m.start() == 0 else 0.0
        score = start_bonus + coverage
        if score > best_score:
            best_score = score
            best_row = row
            best_label = label

    return (best_row, best_label) if best_row is not None else None


def _get_aggregate_value(sd: SheetData, row: int, label_col: int = 2) -> tuple[Any, int | None]:
    """
    Get the most representative value for a KPI row.
    Prefers summary column or last non-empty numeric in row.
    """
    best_val = None
    best_col = None
    summary_col = _find_summary_col(sd)

    if summary_col:
        cd = sd.cells.get((row, summary_col))
        if cd and is_numeric(cd.value):
            return cd.value, summary_col

    # Fallback: first (leftmost) non-empty numeric column.
    # Cap at 200 to guard against broken max_col (~1M on FACT/% sheets).
    for col in range(label_col + 1, min(sd.max_col + 1, 200)):
        cd = sd.cells.get((row, col))
        if cd and is_numeric(cd.value):
            return cd.value, col

    return None, None


def resolve_kpis_preview(
    wb_v1: WorkbookData,
    wb_v2: WorkbookData,
    bd: BusinessDictionary,
    label_col: int = 2,
    wb_v3: WorkbookData | None = None,
) -> list[KPIResolution]:
    """
    Phase-1 preview: resolve KPI row/cell addresses for both workbooks
    using the auto-search algorithm. Returns a list of KPIResolution
    objects for the user to review and correct before the main compare.
    """
    from openpyxl.utils import get_column_letter

    resolutions: list[KPIResolution] = []

    for kpi in bd.kpi_dict:
        res = KPIResolution(
            kpi_name=kpi.name,
            kpi_group=kpi.group,
            kpi_level=kpi.level,
            search_pattern=kpi.search_pattern,
        )

        # Resolve V1
        for sheet_name, sd in wb_v1.sheets.items():
            match = _find_kpi_row(sd, kpi, label_col)
            if match is None:
                continue
            row_idx, label = match
            _, col = _get_aggregate_value(sd, row_idx, label_col)
            col = col or 3
            res.sheet_v1 = sheet_name
            res.row_v1 = row_idx
            res.col_v1 = col
            res.label_v1 = str(label) if label is not None else ""
            res.addr_v1 = f"{sheet_name}!{get_column_letter(col)}{row_idx}"
            res.unit_v1 = _find_unit_in_row(sd, row_idx, label_col) or kpi.unit
            break

        # Resolve V2
        for sheet_name, sd in wb_v2.sheets.items():
            match = _find_kpi_row(sd, kpi, label_col)
            if match is None:
                continue
            row_idx, label = match
            _, col = _get_aggregate_value(sd, row_idx, label_col)
            col = col or 3
            res.sheet_v2 = sheet_name
            res.row_v2 = row_idx
            res.col_v2 = col
            res.label_v2 = str(label) if label is not None else ""
            res.addr_v2 = f"{sheet_name}!{get_column_letter(col)}{row_idx}"
            res.unit_v2 = _find_unit_in_row(sd, row_idx, label_col) or kpi.unit
            break

        # Resolve V3 (optional — Stage 4)
        if wb_v3 is not None:
            for sheet_name, sd in wb_v3.sheets.items():
                match = _find_kpi_row(sd, kpi, label_col)
                if match is None:
                    continue
                row_idx, label = match
                _, col = _get_aggregate_value(sd, row_idx, label_col)
                col = col or 3
                res.sheet_v3 = sheet_name
                res.row_v3 = row_idx
                res.col_v3 = col
                res.label_v3 = str(label) if label is not None else ""
                res.addr_v3 = f"{sheet_name}!{get_column_letter(col)}{row_idx}"
                res.unit_v3 = _find_unit_in_row(sd, row_idx, label_col) or kpi.unit
                break

        resolutions.append(res)

    return resolutions


def _resolution_to_addr(
    res: KPIResolution,
    version: str,
) -> CellAddress | None:
    """Convert a KPIResolution to a CellAddress for the given version ("v1"/"v2"/"v3")."""
    if version == "v1":
        sheet, row, col = res.sheet_v1, res.row_v1, res.col_v1
    elif version == "v3":
        sheet, row, col = res.sheet_v3, res.row_v3, res.col_v3
    else:
        sheet, row, col = res.sheet_v2, res.row_v2, res.col_v2
    if not sheet or row is None or col is None:
        return None
    return CellAddress(sheet=sheet, row=row, col=col)


def extract_kpis(
    wb: WorkbookData,
    bd: BusinessDictionary,
    label_col: int = 2,
    addr_overrides: dict[str, CellAddress] | None = None,
) -> dict[str, tuple[Any, CellAddress | None]]:
    """
    Extract all KPI values from workbook.
    Returns dict: kpi_name → (value, CellAddress).

    If addr_overrides is provided (kpi_name → CellAddress), the override
    address is used directly (reading the cell value from the workbook)
    instead of running the row-search algorithm.
    """
    results: dict[str, tuple[Any, CellAddress | None]] = {}
    found = 0

    for kpi in bd.kpi_dict:
        # Use confirmed override if available
        if addr_overrides and kpi.name in addr_overrides:
            addr = addr_overrides[kpi.name]
            sd = wb.sheets.get(addr.sheet)
            if sd is not None:
                cd = sd.cells.get((addr.row, addr.col))
                value = cd.value if cd else None
                results[kpi.name] = (value, addr)
                if value is not None:
                    found += 1
            else:
                results[kpi.name] = (None, None)
            continue

        for sheet_name, sd in wb.sheets.items():
            match = _find_kpi_row(sd, kpi, label_col)
            if match is None:
                continue
            row_idx, _ = match
            value, col = _get_aggregate_value(sd, row_idx, label_col)
            if value is not None:
                addr = CellAddress(sheet=sheet_name, row=row_idx, col=col or 3)
                results[kpi.name] = (value, addr)
                found += 1
                break
        else:
            if kpi.name not in results:
                results[kpi.name] = (None, None)

    log.info(f"KPI extraction: found={found} total={len(bd.kpi_dict)}")
    return results


def build_kpi_comparison(
    kpi_v1: dict[str, tuple[Any, CellAddress | None]],
    kpi_v2: dict[str, tuple[Any, CellAddress | None]],
    bd: BusinessDictionary,
) -> list[KPIValue]:
    """Compare KPI values between V1 and V2."""
    comparison: list[KPIValue] = []
    kpi_by_name = {k.name: k for k in bd.kpi_dict}

    all_names = list(dict.fromkeys(list(kpi_v1.keys()) + list(kpi_v2.keys())))

    for name in all_names:
        v1_val, addr_v1 = kpi_v1.get(name, (None, None))
        v2_val, addr_v2 = kpi_v2.get(name, (None, None))
        kpi_entry = kpi_by_name.get(name)

        delta: Any = None
        delta_pct: float | None = None
        impact = "neutral"
        note = ""

        if is_numeric(v1_val) and is_numeric(v2_val):
            delta = v2_val - v1_val
            if v1_val != 0:
                delta_pct = (delta / abs(v1_val)) * 100
            else:
                delta_pct = None  # division by zero — no meaningful percentage

            if kpi_entry:
                direction = kpi_entry.better_direction
                if direction == "up":
                    impact = "positive" if delta > 0 else ("negative" if delta < 0 else "neutral")
                elif direction == "down":
                    impact = "positive" if delta < 0 else ("negative" if delta > 0 else "neutral")
        elif v1_val is None and v2_val is None:
            note = "KPI не найден"
            impact = "neutral"
        elif v1_val is None:
            note = "Только в V2"
        elif v2_val is None:
            note = "Только в V1"

        comparison.append(KPIValue(
            kpi_name=name,
            kpi_group=kpi_entry.group if kpi_entry else "Прочее",
            kpi_level=kpi_entry.level if kpi_entry else 2,
            unit=kpi_entry.unit if kpi_entry else "",
            value_v1=v1_val,
            value_v2=v2_val,
            delta=delta,
            delta_pct=delta_pct,
            impact=impact,
            note=note,
            addr_v1=addr_v1,
            addr_v2=addr_v2,
        ))

    return comparison
