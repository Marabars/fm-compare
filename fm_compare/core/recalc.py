"""
Formula recalculation engine via LibreOffice headless.

recalc_with_overrides() copies the source file, patches raw cell values
from the `overrides` dict, then runs `soffice --headless --convert-to xlsx`
so LibreOffice recalculates all dependent formulas.  The resulting file is
read back with the standard excel_reader and returned as WorkbookData.

RecalcError is raised when LibreOffice is missing, times out, or exits
with a non-zero code.  Callers that want graceful degradation should catch
RecalcError and fall back to reading the original file without recalculation.
"""
from __future__ import annotations

import re
import shutil
import subprocess
import tempfile
from pathlib import Path
from typing import Any

from openpyxl.utils import column_index_from_string, get_column_letter

from fm_compare.core.excel_reader import load_workbook_data, WorkbookData
from fm_compare.core.models import CellAddress


class RecalcError(RuntimeError):
    pass


# ── LibreOffice discovery ────────────────────────────────────────────────────

_SOFFICE_CANDIDATES = (
    "soffice",
    "libreoffice",
    "/usr/bin/soffice",
    "/usr/lib/libreoffice/program/soffice",
    "/opt/libreoffice/program/soffice",
)


def _find_soffice() -> str:
    for candidate in _SOFFICE_CANDIDATES:
        if shutil.which(candidate):
            return candidate
    raise RecalcError(
        "LibreOffice not found — install libreoffice-calc on this system"
    )


# ── Public API ───────────────────────────────────────────────────────────────

def recalc_with_overrides(
    path: Path,
    overrides: dict[str, Any],
    timeout: int = 90,
) -> WorkbookData:
    """
    Copy `path` to a temp dir, patch cells from `overrides`, recalculate with
    LibreOffice headless, return WorkbookData with fresh computed values.

    `overrides` keys must be strings in "Sheet!A1" format.
    Use addr_to_str(CellAddress) to convert before calling.
    """
    soffice = _find_soffice()
    tmp = Path(tempfile.mkdtemp(prefix="fm_recalc_"))
    try:
        src = tmp / path.name
        shutil.copy2(path, src)
        _patch_overrides(src, overrides)
        out_dir = tmp / "out"
        out_dir.mkdir()
        _run_libreoffice(soffice, src, out_dir, timeout)
        out_file = _find_output(out_dir, src.stem)
        return load_workbook_data(out_file)
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def addr_to_str(addr: CellAddress) -> str:
    """Convert CellAddress → "Sheet!B5" string (for dict keys)."""
    return f"{addr.sheet}!{get_column_letter(addr.col)}{addr.row}"


# ── Internal helpers ─────────────────────────────────────────────────────────

def _patch_overrides(path: Path, overrides: dict[str, Any]) -> None:
    import openpyxl
    wb = openpyxl.load_workbook(str(path), data_only=False)
    for key, value in overrides.items():
        sheet_name, row, col = _parse_addr_str(str(key))
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        # Replace formula/value so LibreOffice sees a plain number.
        ws.cell(row=row, column=col).value = value
    wb.save(str(path))


def _parse_addr_str(addr: str) -> tuple[str, int, int]:
    """Parse "Sheet!A1" → (sheet, row, col)."""
    if "!" not in addr:
        raise RecalcError(f"Expected Sheet!Cell format, got: {addr!r}")
    sheet, cell = addr.split("!", 1)
    m = re.match(r"^([A-Z]+)(\d+)$", cell.upper().strip())
    if not m:
        raise RecalcError(f"Cannot parse cell address: {cell!r}")
    col = column_index_from_string(m.group(1))
    row = int(m.group(2))
    return sheet, row, col


def _run_libreoffice(soffice: str, src: Path, out_dir: Path, timeout: int) -> None:
    cmd = [
        soffice,
        "--headless",
        "--norestore",
        "--invisible",
        "--convert-to", "xlsx",
        "--outdir", str(out_dir),
        str(src),
    ]
    try:
        result = subprocess.run(
            cmd, capture_output=True, text=True, timeout=timeout
        )
        if result.returncode != 0:
            raise RecalcError(
                f"LibreOffice exited {result.returncode}: {result.stderr[:400]}"
            )
    except subprocess.TimeoutExpired as exc:
        raise RecalcError(f"LibreOffice timed out after {timeout}s") from exc


def _find_output(out_dir: Path, stem: str) -> Path:
    """Locate the recalculated file; LO may emit <stem>.xlsx regardless of input ext."""
    for ext in (".xlsx", ".xlsm"):
        p = out_dir / (stem + ext)
        if p.exists():
            return p
    # Fallback: any xlsx in the dir
    candidates = list(out_dir.glob("*.xlsx")) + list(out_dir.glob("*.xlsm"))
    if candidates:
        return candidates[0]
    raise RecalcError(f"LibreOffice produced no output file in {out_dir}")
