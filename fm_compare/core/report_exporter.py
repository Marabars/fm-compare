"""
Excel report exporter. Generates the output .xlsx with all required sheets.
Sheet names in English, content in Russian.
"""
from __future__ import annotations
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, Rule
from openpyxl.utils import get_column_letter

from fm_compare.core.models import (
    CompareResult, CompareMode, KPIValue, DiffRow, FormulaChange,
    TimingShift, Warning, Severity, ChangeType, MatchType
)
from fm_compare.core.business_dictionary import BusinessDictionary
from fm_compare.security import safe_logger as log


# Color palette
C_GREEN = "C6EFCE"
C_RED = "FFC7CE"
C_YELLOW = "FFEB9C"
C_GRAY = "D9D9D9"
C_BLUE_HEADER = "4472C4"
C_WHITE = "FFFFFF"
C_LIGHT_BLUE = "DCE6F1"

FONT_HEADER = Font(bold=True, color=C_WHITE, size=10)
FONT_BOLD = Font(bold=True, size=10)
FONT_NORMAL = Font(size=10)
FILL_HEADER = PatternFill("solid", fgColor=C_BLUE_HEADER)
FILL_SUBHEADER = PatternFill("solid", fgColor=C_LIGHT_BLUE)
FILL_GREEN = PatternFill("solid", fgColor=C_GREEN)
FILL_RED = PatternFill("solid", fgColor=C_RED)
FILL_YELLOW = PatternFill("solid", fgColor=C_YELLOW)
FILL_GRAY = PatternFill("solid", fgColor=C_GRAY)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)


def _h(ws, row: int, col: int, value: Any, subheader: bool = False) -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.font = FONT_HEADER if not subheader else Font(bold=True, size=10)
    c.fill = FILL_HEADER if not subheader else FILL_SUBHEADER
    c.alignment = ALIGN_CENTER
    c.border = THIN_BORDER


def _cell(ws, row: int, col: int, value: Any, fill: PatternFill | None = None,
          bold: bool = False, align: str = "left", number_format: str = "General") -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, size=10)
    if fill:
        c.fill = fill
    c.alignment = Alignment(
        horizontal=align, vertical="top", wrap_text=(align == "left")
    )
    c.number_format = number_format
    c.border = THIN_BORDER


def _auto_width(ws, min_w: int = 8, max_w: int = 50) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_w), max_w)


def _freeze(ws, cell: str = "A2") -> None:
    ws.freeze_panes = cell


def export_report(
    result: CompareResult,
    bd: BusinessDictionary,
    output_path: Path,
    mode: CompareMode = CompareMode.FULL,
) -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    _add_executive_summary(wb, result, mode)
    _add_kpi_comparison(wb, result.kpi_values)
    _add_top_changes(wb, result.diff_rows, result.run_settings.get("top_x", 10))
    _add_business_diff(wb, result.diff_rows, mode)
    _add_formula_changes(wb, result.formula_changes)
    _add_timing_shifts(wb, result.timing_shifts)

    # Hidden sheets
    _add_comments_changes(wb, result.comment_changes)
    _add_hidden_rows_changes(wb, result.hidden_row_changes)
    _add_warnings(wb, result.warnings)

    if mode == CompareMode.FULL:
        _add_raw_diff(wb, result.raw_diff_rows)

    _add_run_settings(wb, result.run_settings, mode)
    _add_dictionary_export(wb, bd)

    # Set Executive Summary as first active sheet
    wb.active = wb["Executive Summary"]

    # Write to temp first, then move to final path
    tmp = output_path.with_suffix(".tmp.xlsx")
    wb.save(str(tmp))
    if output_path.exists():
        output_path.unlink()
    tmp.rename(output_path)

    log.info(f"Report exported ({output_path.stat().st_size // 1024} KB)")
    return output_path


def _add_executive_summary(wb: Workbook, result: CompareResult, mode: CompareMode) -> None:
    ws = wb.create_sheet("Executive Summary")
    row = 1

    # Title
    ws.merge_cells(f"A{row}:H{row}")
    c = ws.cell(row=row, column=1, value="Сравнение версий финансовой модели — Executive Summary")
    c.font = Font(bold=True, size=14, color=C_WHITE)
    c.fill = FILL_HEADER
    c.alignment = ALIGN_CENTER
    row += 1

    run_date = result.run_settings.get("run_date", datetime.now().strftime("%Y-%m-%d %H:%M"))
    ws.merge_cells(f"A{row}:H{row}")
    c = ws.cell(row=row, column=1, value=f"Дата: {run_date}  |  Режим: {'Полный аудит' if mode == CompareMode.FULL else 'Quick KPI Check'}")
    c.font = Font(size=10, italic=True)
    c.alignment = ALIGN_CENTER
    row += 2

    for block in result.summary_blocks:
        btype = block.get("type", "")

        # Block title
        ws.merge_cells(f"A{row}:H{row}")
        title_fill = FILL_YELLOW if "warning" in btype or "sign" in btype else FILL_SUBHEADER
        c = ws.cell(row=row, column=1, value=block.get("title", ""))
        c.font = Font(bold=True, size=11)
        c.fill = title_fill
        c.alignment = ALIGN_LEFT
        row += 1

        # Block text
        if block.get("text"):
            ws.merge_cells(f"A{row}:H{row}")
            c = ws.cell(row=row, column=1, value=block["text"])
            c.font = FONT_NORMAL
            c.alignment = ALIGN_LEFT
            ws.row_dimensions[row].height = 30
            row += 1

        # Items
        for item in block.get("items", []):
            ws.merge_cells(f"A{row}:H{row}")
            c = ws.cell(row=row, column=1, value=str(item))
            c.font = FONT_NORMAL
            c.alignment = ALIGN_LEFT
            if str(item).startswith("▼"):
                c.fill = FILL_RED
            elif str(item).startswith("▲"):
                c.fill = FILL_GREEN
            elif "⚠" in str(item) or "не найден" in str(item).lower():
                c.fill = FILL_YELLOW
            row += 1

        row += 1  # blank row between blocks

    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 20
    ws.column_dimensions["A"].width = 80
    _freeze(ws, "A3")


def _add_kpi_comparison(wb: Workbook, kpi_values: list[KPIValue]) -> None:
    ws = wb.create_sheet("KPI Comparison")
    headers = ["Группа KPI", "Уровень", "KPI", "Ед.", "V1", "V2", "Δ", "Δ%", "Impact", "Примечание"]
    for ci, h in enumerate(headers, 1):
        _h(ws, 1, ci, h)

    for ri, k in enumerate(kpi_values, 2):
        fill = None
        if k.impact == "positive":
            fill = FILL_GREEN
        elif k.impact == "negative":
            fill = FILL_RED
        elif k.value_v1 is None and k.value_v2 is None:
            fill = FILL_YELLOW

        _cell(ws, ri, 1, k.kpi_group, fill, align="center")
        _cell(ws, ri, 2, k.kpi_level, fill, align="center")
        _cell(ws, ri, 3, k.kpi_name, fill, bold=(k.kpi_level == 1))
        _cell(ws, ri, 4, k.unit, fill, align="center")
        _cell(ws, ri, 5, k.value_v1, fill, align="right", number_format="#,##0.00")
        _cell(ws, ri, 6, k.value_v2, fill, align="right", number_format="#,##0.00")
        _cell(ws, ri, 7, k.delta, fill, align="right", number_format="+#,##0.00;-#,##0.00")
        pct = f"{k.delta_pct:+.1f}%" if k.delta_pct is not None and abs(k.delta_pct) < 10000 else ""
        _cell(ws, ri, 8, pct, fill, align="center")
        _cell(ws, ri, 9, k.impact, fill, align="center")
        _cell(ws, ri, 10, k.note, fill)

    _auto_width(ws)
    _freeze(ws, "A2")
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _add_top_changes(wb: Workbook, diff_rows: list[DiffRow], top_x: int = 10) -> None:
    ws = wb.create_sheet("Top Changes")
    headers = ["Лист", "Бизнес-ключ V1", "Бизнес-ключ V2", "V1", "V2", "Δ", "Δ%", "Направление"]
    for ci, h in enumerate(headers, 1):
        _h(ws, 1, ci, h)

    material = [
        d for d in diff_rows
        if d.is_material and isinstance(d.delta, (int, float)) and d.delta != 0
    ]
    material.sort(key=lambda d: abs(d.delta) if d.delta else 0, reverse=True)

    half = top_x // 2
    extra = top_x % 2
    neg = [d for d in material if d.delta < 0][:half + extra]
    pos = [d for d in material if d.delta > 0][:half]
    top = neg + pos

    for ri, d in enumerate(top, 2):
        fill = FILL_RED if d.delta < 0 else FILL_GREEN
        sheet = d.addr_v1.sheet if d.addr_v1 else (d.addr_v2.sheet if d.addr_v2 else "")
        _cell(ws, ri, 1, sheet, fill, align="center")
        _cell(ws, ri, 2, d.business_key_v1, fill)
        _cell(ws, ri, 3, d.business_key_v2, fill)
        _cell(ws, ri, 4, d.value_v1, fill, align="right", number_format="#,##0.00")
        _cell(ws, ri, 5, d.value_v2, fill, align="right", number_format="#,##0.00")
        _cell(ws, ri, 6, d.delta, fill, align="right", number_format="+#,##0.00;-#,##0.00")
        pct = f"{d.delta_pct:+.1f}%" if d.delta_pct is not None else ""
        _cell(ws, ri, 7, pct, fill, align="center")
        _cell(ws, ri, 8, "▼ Ухудшение" if d.delta < 0 else "▲ Улучшение", fill, align="center")

    _auto_width(ws)
    _freeze(ws, "A2")


def _add_business_diff(wb: Workbook, diff_rows: list[DiffRow], mode: CompareMode) -> None:
    ws = wb.create_sheet("Business Diff")
    headers = [
        "Лист", "Бизнес-ключ V1", "Бизнес-ключ V2",
        "Тип сопоставления", "Уверенность",
        "Адрес V1", "Адрес V2",
        "V1", "V2", "Δ", "Δ%",
        "Тип изменения", "Существенное", "Знак изменён",
        "Группа KPI", "Формула V1", "Формула V2"
    ]
    for ci, h in enumerate(headers, 1):
        _h(ws, 1, ci, h)

    rows_to_show = (
        diff_rows if mode == CompareMode.FULL
        else [d for d in diff_rows if d.is_material]
    )

    for ri, d in enumerate(rows_to_show, 2):
        fill = None
        if d.sign_changed:
            fill = PatternFill("solid", fgColor="FF0000")
        elif d.change_type == ChangeType.NEW_ITEM:
            fill = FILL_GREEN
        elif d.change_type == ChangeType.DELETED_ITEM:
            fill = FILL_RED
        elif d.match_type == MatchType.FUZZY:
            fill = FILL_YELLOW
        elif d.match_type == MatchType.NOT_MATCHED:
            fill = FILL_YELLOW

        addr_v1_str = str(d.addr_v1) if d.addr_v1 else ""
        addr_v2_str = str(d.addr_v2) if d.addr_v2 else ""

        vals = [
            d.addr_v1.sheet if d.addr_v1 else "",
            d.business_key_v1, d.business_key_v2,
            d.match_type.value, f"{d.match_confidence:.0%}",
            addr_v1_str, addr_v2_str,
            d.value_v1, d.value_v2, d.delta,
            f"{d.delta_pct:+.1f}%" if d.delta_pct else "",
            d.change_type.value,
            "Да" if d.is_material else "Нет",
            "⚠ Да" if d.sign_changed else "Нет",
            d.kpi_group or "",
            d.formula_v1 or "", d.formula_v2 or "",
        ]
        for ci, v in enumerate(vals, 1):
            _cell(ws, ri, ci, v, fill, number_format="#,##0.00" if ci in (8, 9, 10) else "General")

    _auto_width(ws)
    _freeze(ws, "A2")
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _add_formula_changes(wb: Workbook, changes: list[FormulaChange]) -> None:
    ws = wb.create_sheet("Formula Changes")
    headers = [
        "Лист", "Адрес V1", "Адрес V2",
        "Формула V1", "Формула V2",
        "Значение V1", "Значение V2",
        "Только логика", "Частичная трассировка", "Зависимые KPI"
    ]
    for ci, h in enumerate(headers, 1):
        _h(ws, 1, ci, h)

    for ri, fc in enumerate(changes, 2):
        fill = FILL_YELLOW if fc.logic_changed else None
        _cell(ws, ri, 1, fc.addr_v1.sheet if fc.addr_v1 else "")
        _cell(ws, ri, 2, str(fc.addr_v1) if fc.addr_v1 else "", fill)
        _cell(ws, ri, 3, str(fc.addr_v2) if fc.addr_v2 else "", fill)
        _cell(ws, ri, 4, fc.formula_v1, fill)
        _cell(ws, ri, 5, fc.formula_v2, fill)
        _cell(ws, ri, 6, fc.value_v1, fill, number_format="#,##0.00")
        _cell(ws, ri, 7, fc.value_v2, fill, number_format="#,##0.00")
        _cell(ws, ri, 8, "⚠ Да" if fc.logic_changed else "Нет", fill, align="center")
        _cell(ws, ri, 9, "Да" if fc.dependency_partial else "Нет", fill, align="center")
        _cell(ws, ri, 10, ", ".join(fc.affected_kpi) if fc.affected_kpi else "", fill)

    _auto_width(ws)
    _freeze(ws, "A2")
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _add_timing_shifts(wb: Workbook, shifts: list[TimingShift]) -> None:
    ws = wb.create_sheet("Timing Shifts")
    headers = ["Лист", "Группа KPI", "Бизнес-ключ", "Сдвиг (периодов)", "Объём сдвига", "Адрес V1", "Адрес V2"]
    for ci, h in enumerate(headers, 1):
        _h(ws, 1, ci, h)

    for ri, s in enumerate(shifts, 2):
        fill = FILL_RED if abs(s.periods_shifted) >= 3 else FILL_YELLOW
        _cell(ws, ri, 1, s.sheet, fill, align="center")
        _cell(ws, ri, 2, s.kpi_group, fill)
        _cell(ws, ri, 3, s.business_key, fill)
        _cell(ws, ri, 4, s.periods_shifted, fill, align="center")
        _cell(ws, ri, 5, s.amount_shifted, fill, number_format="#,##0.00")
        _cell(ws, ri, 6, str(s.addr_v1) if s.addr_v1 else "", fill)
        _cell(ws, ri, 7, str(s.addr_v2) if s.addr_v2 else "", fill)

    _auto_width(ws)
    _freeze(ws, "A2")


def _add_comments_changes(wb: Workbook, comment_changes: list[dict]) -> None:
    ws = wb.create_sheet("Comments Changes")
    ws.sheet_state = "hidden"
    headers = ["Лист", "Адрес V1", "Адрес V2", "Комментарий V1", "Комментарий V2"]
    for ci, h in enumerate(headers, 1):
        _h(ws, 1, ci, h)
    for ri, cc in enumerate(comment_changes, 2):
        for ci, k in enumerate(["sheet", "addr_v1", "addr_v2", "comment_v1", "comment_v2"], 1):
            _cell(ws, ri, ci, cc.get(k, ""))
    _auto_width(ws)


def _add_hidden_rows_changes(wb: Workbook, hidden_changes: list[dict]) -> None:
    ws = wb.create_sheet("Hidden Rows Changes")
    ws.sheet_state = "hidden"
    headers = ["Лист", "Строка V1", "Строка V2", "Скрыта V1", "Скрыта V2"]
    for ci, h in enumerate(headers, 1):
        _h(ws, 1, ci, h)
    for ri, hc in enumerate(hidden_changes, 2):
        for ci, k in enumerate(["sheet", "row_v1", "row_v2", "hidden_v1", "hidden_v2"], 1):
            _cell(ws, ri, ci, hc.get(k, ""))
    _auto_width(ws)


def _add_warnings(wb: Workbook, warnings: list[Warning]) -> None:
    ws = wb.create_sheet("Warnings")
    ws.sheet_state = "hidden"
    headers = ["Severity", "Категория", "Сообщение", "Лист", "KPI", "Ручная проверка"]
    for ci, h in enumerate(headers, 1):
        _h(ws, 1, ci, h)

    severity_fills = {
        Severity.CRITICAL: PatternFill("solid", fgColor="FF0000"),
        Severity.HIGH: FILL_RED,
        Severity.MEDIUM: FILL_YELLOW,
        Severity.LOW: FILL_GRAY,
        Severity.INFO: None,
    }
    for ri, w in enumerate(warnings, 2):
        fill = severity_fills.get(w.severity)
        _cell(ws, ri, 1, w.severity.value, fill, align="center")
        _cell(ws, ri, 2, w.category, fill)
        _cell(ws, ri, 3, w.message, fill)
        _cell(ws, ri, 4, w.related_sheet, fill)
        _cell(ws, ri, 5, w.related_kpi, fill)
        _cell(ws, ri, 6, "Да" if w.manual_check_required else "Нет", fill, align="center")

    _auto_width(ws)
    _freeze(ws, "A2")
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _add_raw_diff(wb: Workbook, raw_rows: list[dict]) -> None:
    ws = wb.create_sheet("Raw Diff")
    ws.sheet_state = "hidden"
    if not raw_rows:
        ws.cell(row=1, column=1, value="Raw Diff пуст")
        return

    headers = list(raw_rows[0].keys()) if raw_rows else []
    for ci, h in enumerate(headers, 1):
        _h(ws, 1, ci, h)
    for ri, row in enumerate(raw_rows, 2):
        for ci, k in enumerate(headers, 1):
            _cell(ws, ri, ci, row.get(k, ""))
    _auto_width(ws)
    _freeze(ws, "A2")


def _add_run_settings(wb: Workbook, settings: dict, mode: CompareMode) -> None:
    ws = wb.create_sheet("Run Settings")
    ws.sheet_state = "hidden"

    ws.merge_cells("A1:C1")
    c = ws.cell(row=1, column=1, value="Настройки запуска сравнения")
    c.font = Font(bold=True, size=12)
    c.fill = FILL_HEADER
    c.font = Font(bold=True, color=C_WHITE, size=12)
    c.alignment = ALIGN_CENTER

    row = 3
    ws.cell(row=row, column=1, value="КОНФИДЕНЦИАЛЬНО — только для внутреннего использования")
    ws.cell(row=row, column=1).font = Font(bold=True, color="FF0000")
    row += 1
    ws.cell(row=row, column=1, value="Финансовые данные обрабатываются локально и не покидают доверенный контур.")
    row += 2

    display_settings = {
        "Режим": "Полный аудит" if mode == CompareMode.FULL else "Quick KPI Check",
        "Дата и время": settings.get("run_date", ""),
        "Версия приложения": settings.get("app_version", ""),
        "Порог материальности (абс.)": settings.get("materiality_abs", "не задан"),
        "Порог материальности (%)": settings.get("materiality_pct", "не задан"),
        "Top-X изменений": settings.get("top_x", 10),
        "Листов V1 выбрано": settings.get("sheets_v1_count", ""),
        "Листов V2 выбрано": settings.get("sheets_v2_count", ""),
        "Включены комментарии": settings.get("include_comments", True),
        "Включены скрытые строки": settings.get("include_hidden_rows", True),
    }
    for k, v in display_settings.items():
        _cell(ws, row, 1, k, bold=True)
        _cell(ws, row, 2, str(v))
        row += 1

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 30


def _add_dictionary_export(wb: Workbook, bd: BusinessDictionary) -> None:
    ws = wb.create_sheet("Dictionary Export")
    ws.sheet_state = "hidden"

    row = 1
    ws.cell(row=row, column=1, value="Использованный бизнес-словарь")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 2

    # KPI list
    ws.cell(row=row, column=1, value="KPI Dictionary").font = Font(bold=True)
    row += 1
    for ci, h in enumerate(["Название", "Группа", "Уровень", "Ед.", "Направление улучшения"], 1):
        _h(ws, row, ci, h)
    row += 1
    for k in bd.kpi_dict:
        ws.cell(row=row, column=1, value=k.name)
        ws.cell(row=row, column=2, value=k.group)
        ws.cell(row=row, column=3, value=k.level)
        ws.cell(row=row, column=4, value=k.unit)
        ws.cell(row=row, column=5, value=k.better_direction)
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="Sheet Dictionary").font = Font(bold=True)
    row += 1
    for ci, h in enumerate(["Паттерн", "Группа", "Ключевой лист", "Анализировать по умолч."], 1):
        _h(ws, row, ci, h)
    row += 1
    for s in bd.sheet_dict:
        ws.cell(row=row, column=1, value=s.name_pattern)
        ws.cell(row=row, column=2, value=s.group)
        ws.cell(row=row, column=3, value="Да" if s.key_sheet else "Нет")
        ws.cell(row=row, column=4, value="Да" if s.analyze_by_default else "Нет")
        row += 1

    _auto_width(ws)


def suggest_filename(mode: CompareMode) -> str:
    ts = datetime.now().strftime("%Y-%m-%d_%H%M")
    suffix = "Full" if mode == CompareMode.FULL else "Quick"
    return f"Model_Comparison_{suffix}_{ts}.xlsx"
