"""
Excel reader: loads .xlsx, extracts cell values, formulas, comments,
hidden rows. Never logs business data.
"""
from __future__ import annotations
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet

from fm_compare.core.models import CellAddress, WorkbookInfo, PreCheckResult
from fm_compare.security import safe_logger as log


SIZE_WARNING_MB = 30.0


@dataclass
class CellData:
    address: CellAddress
    value: Any
    formula: str | None              # raw formula string e.g. "=SUM(A1:A10)"
    data_type: str                   # "n", "s", "b", "e", "str", "inlineStr", "d"
    comment: str | None
    is_hidden_row: bool


@dataclass
class SheetData:
    name: str
    cells: dict[tuple[int, int], CellData] = field(default_factory=dict)
    row_hidden: dict[int, bool] = field(default_factory=dict)
    col_widths: dict[int, float] = field(default_factory=dict)
    max_row: int = 0
    max_col: int = 0
    period_headers: list[tuple[int, int, Any]] = field(default_factory=list)


@dataclass
class WorkbookData:
    info: WorkbookInfo = field(default_factory=WorkbookInfo)
    sheets: dict[str, SheetData] = field(default_factory=dict)


def _safe_value(cell: Any) -> Any:
    """Returns cached value from a cell. Returns None for EmptyCell."""
    v = getattr(cell, "value", None)
    if v is None:
        return None
    dt = getattr(cell, "data_type", None)
    if dt == "e":
        return f"#ERROR:{v}"
    return v


def _formula_str(cell: Any) -> str | None:
    # Only trust data_type == "f"; startswith("=") misidentifies literal text cells.
    dt = getattr(cell, "data_type", None)
    if dt == "f":
        v = getattr(cell, "value", None)
        return str(v) if v is not None else None
    return None


def load_workbook_data(
    path: Path,
    selected_sheets: list[str] | None = None,
) -> WorkbookData:
    """
    Load workbook from path. Reads cached values (data_only=True) for values
    and also reads formulas (data_only=False) for formula comparison.
    """
    log.info(f"Loading workbook ({path.stat().st_size // 1024} KB)")

    wb_values: Workbook = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    wb_formulas: Workbook = openpyxl.load_workbook(str(path), data_only=False, read_only=True)

    info = WorkbookInfo()
    info.sheet_names = wb_values.sheetnames
    info.visible_sheets = [
        s for s in wb_values.sheetnames
        if wb_values[s].sheet_state == "visible"
    ]
    info.hidden_sheets = [
        s for s in wb_values.sheetnames
        if wb_values[s].sheet_state != "visible"
    ]
    info.file_size_mb = path.stat().st_size / (1024 * 1024)

    target_sheets = selected_sheets if selected_sheets else info.sheet_names
    result = WorkbookData(info=info)

    for sheet_name in target_sheets:
        if sheet_name not in wb_values.sheetnames:
            log.warning(f"Sheet not found in workbook (name hash {hash(sheet_name) & 0xFFFF})")
            continue

        ws_val: Worksheet = wb_values[sheet_name]
        ws_fml: Worksheet = wb_formulas[sheet_name]

        sd = SheetData(name=sheet_name)
        sd.max_row = ws_val.max_row or 0
        sd.max_col = ws_val.max_column or 0

        # read_only=True does not populate row_dimensions — hidden rows won't be detected.
        if hasattr(ws_val, "row_dimensions"):
            for row_idx, rd in ws_val.row_dimensions.items():
                sd.row_hidden[row_idx] = rd.hidden or False

        # Skip iter_rows on empty sheets — avoids TypeError when max_row is None.
        if sd.max_row == 0 or sd.max_col == 0:
            result.sheets[sheet_name] = sd
            log.info("Sheet loaded: 0 cells (empty)")
            continue

        val_rows = list(ws_val.iter_rows())
        fml_rows = list(ws_fml.iter_rows())

        if len(val_rows) != len(fml_rows):
            log.warning(
                f"Val/formula row count mismatch (val={len(val_rows)}, fml={len(fml_rows)})"
            )

        for r_idx, (val_row, fml_row) in enumerate(zip(val_rows, fml_rows), start=1):
            for c_idx, (vc, fc) in enumerate(zip(val_row, fml_row), start=1):
                value = _safe_value(vc)
                formula = _formula_str(fc)
                comment = (vc.comment.text if vc.comment else None) if hasattr(vc, "comment") else None
                hidden = sd.row_hidden.get(r_idx, False)

                if value is None and formula is None and comment is None:
                    continue

                addr = CellAddress(sheet=sheet_name, row=r_idx, col=c_idx)
                cd = CellData(
                    address=addr,
                    value=value,
                    formula=formula,
                    data_type=getattr(vc, "data_type", None) or "n",
                    comment=comment,
                    is_hidden_row=hidden,
                )
                sd.cells[(r_idx, c_idx)] = cd

        result.sheets[sheet_name] = sd
        log.info(f"Sheet loaded: {len(sd.cells)} cells")

    wb_values.close()
    wb_formulas.close()
    log.info("Workbook load complete")
    return result


def pre_check(path: Path | None, label: str = "") -> PreCheckResult:
    result = PreCheckResult()
    lbl = label or "file"

    if path is None:
        result.ok = False
        result.blocking_errors.append(f"{lbl}: файл не выбран")
        return result

    if not path.exists():
        result.ok = False
        result.blocking_errors.append(f"{lbl}: файл не найден")
        return result

    if path.suffix.lower() != ".xlsx":
        result.ok = False
        result.blocking_errors.append(f"{lbl}: неподдерживаемый формат (только .xlsx)")
        return result

    size_mb = path.stat().st_size / (1024 * 1024)
    if size_mb > SIZE_WARNING_MB:
        result.warnings.append(
            f"{lbl}: файл больше {SIZE_WARNING_MB} МБ — возможна медленная работа"
        )

    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        wb.close()
    except Exception as e:
        result.ok = False
        result.blocking_errors.append(f"{lbl}: не удаётся открыть файл ({type(e).__name__})")

    return result


def get_workbook_info(path: Path) -> WorkbookInfo:
    """Return lightweight WorkbookInfo (sheet names, file size) without loading cell data."""
    names = get_sheet_names(path)
    size_mb = path.stat().st_size / 1_048_576 if path.exists() else 0.0
    return WorkbookInfo(
        sheet_names=names,
        file_size_mb=round(size_mb, 2),
    )


def get_sheet_names(path: Path) -> list[str]:
    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        names = wb.sheetnames
        wb.close()
        return names
    except Exception as e:
        log.warning(f"get_sheet_names failed ({type(e).__name__})")
        return []


def find_period_headers(sd: SheetData, header_rows: int = 3) -> list[tuple[int, int, Any]]:
    """
    Detect period columns by scanning top rows for date-like or year-like headers.
    Returns list of (row, col, value).
    """
    headers = []
    for row in range(1, min(header_rows + 1, sd.max_row + 1)):
        for col in range(1, sd.max_col + 1):
            cd = sd.cells.get((row, col))
            if cd is None:
                continue
            v = cd.value
            if v is None:
                continue
            from datetime import date, datetime as dt
            if isinstance(v, (date, dt)):
                headers.append((row, col, v))
            elif isinstance(v, (int, float)) and not isinstance(v, bool) and 2000 <= v <= 2100:
                headers.append((row, col, int(v)))
            elif isinstance(v, str) and re.match(r"(20\d\d|Q[1-4]\s*20\d\d|\d{4}-\d{2})", v.strip()):
                headers.append((row, col, v.strip()))
    return headers
