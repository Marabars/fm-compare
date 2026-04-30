"""
Business dictionary: loads, validates, stores dictionary configuration.
Handles Sheet Dictionary, KPI Dictionary, Business Key Rules, Materiality Rules,
Synonyms, Warning Rules, Interpretation Rules, Summary Rules, Dependency Rules.
"""
from __future__ import annotations
import json
import os
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from fm_compare.security import safe_logger as log


_DICT_DIR = Path(os.environ.get("APPDATA", Path.home())) / "FM_Compare"
_DICT_FILE = _DICT_DIR / "business_dictionary.json"
_DEFAULT_FILE = Path(__file__).parent.parent / "data" / "default_dictionary.json"


@dataclass
class SheetEntry:
    name_pattern: str
    group: str
    role: str
    key_sheet: bool = False
    analyze_by_default: bool = True
    ignore_business_diff: bool = False


@dataclass
class KPIEntry:
    name: str
    group: str
    level: int           # 1 or 2
    unit: str
    better_direction: str  # "up" | "down" | "neutral"
    search_pattern: str
    search_in: str = "row_label"


@dataclass
class BusinessDictionary:
    sheet_dict: list[SheetEntry] = field(default_factory=list)
    kpi_dict: list[KPIEntry] = field(default_factory=list)
    business_key_rules: list[dict] = field(default_factory=list)
    materiality_rules: dict[str, dict] = field(default_factory=dict)
    synonyms: dict[str, list[str]] = field(default_factory=dict)
    warning_rules: list[dict] = field(default_factory=list)
    interpretation_rules: list[dict] = field(default_factory=list)
    summary_rules: list[dict] = field(default_factory=list)
    dependency_rules: list[dict] = field(default_factory=list)


def _load_raw(path: Path) -> dict:
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def _parse(raw: dict) -> BusinessDictionary:
    bd = BusinessDictionary()
    for s in raw.get("sheet_dictionary", []):
        bd.sheet_dict.append(SheetEntry(**{
            k: s[k] for k in SheetEntry.__dataclass_fields__ if k in s
        }))
    for k in raw.get("kpi_dictionary", []):
        bd.kpi_dict.append(KPIEntry(**{
            f: k[f] for f in KPIEntry.__dataclass_fields__ if f in k
        }))
    bd.business_key_rules = raw.get("business_key_rules", [])
    bd.materiality_rules = raw.get("materiality_rules", {})
    bd.synonyms = raw.get("synonyms", {})
    bd.warning_rules = raw.get("warning_rules", [])
    bd.interpretation_rules = raw.get("interpretation_rules", [])
    bd.summary_rules = raw.get("summary_rules", [])
    bd.dependency_rules = raw.get("dependency_rules", [])
    return bd


def load_dictionary(force_default: bool = False) -> BusinessDictionary:
    if not force_default and _DICT_FILE.exists():
        try:
            raw = _load_raw(_DICT_FILE)
            log.info("Business dictionary loaded from user profile")
            return _parse(raw)
        except Exception as e:
            log.warning(f"Failed to load user dictionary ({type(e).__name__}), using default")

    try:
        raw = _load_raw(_DEFAULT_FILE)
        log.info("Business dictionary loaded from defaults")
        return _parse(raw)
    except Exception as e:
        log.error(f"Failed to load default dictionary ({type(e).__name__})")
        return BusinessDictionary()


def save_dictionary(bd: BusinessDictionary) -> None:
    try:
        _DICT_DIR.mkdir(parents=True, exist_ok=True)
        raw = {
            "sheet_dictionary": [
                {k: getattr(s, k) for k in SheetEntry.__dataclass_fields__}
                for s in bd.sheet_dict
            ],
            "kpi_dictionary": [
                {k: getattr(e, k) for k in KPIEntry.__dataclass_fields__}
                for e in bd.kpi_dict
            ],
            "business_key_rules": bd.business_key_rules,
            "materiality_rules": bd.materiality_rules,
            "synonyms": bd.synonyms,
            "warning_rules": bd.warning_rules,
            "interpretation_rules": bd.interpretation_rules,
            "summary_rules": bd.summary_rules,
            "dependency_rules": bd.dependency_rules,
        }
        with open(_DICT_FILE, "w", encoding="utf-8") as f:
            json.dump(raw, f, ensure_ascii=False, indent=2)
        log.info("Business dictionary saved")
    except Exception as e:
        log.error(f"Failed to save dictionary ({type(e).__name__})")


def reset_to_default() -> BusinessDictionary:
    if _DICT_FILE.exists():
        _DICT_FILE.unlink(missing_ok=True)
    return load_dictionary()


def export_to_excel(bd: BusinessDictionary, path: Path) -> None:
    """Export dictionary to Excel file for editing."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font_white = Font(bold=True, color="FFFFFF")

    def write_sheet(ws, headers: list[str], rows: list[list]) -> None:
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = header_font_white
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center")
        for ri, row in enumerate(rows, 2):
            for ci, val in enumerate(row, 1):
                ws.cell(row=ri, column=ci, value=val)
        ws.freeze_panes = "A2"

    # Sheet Dictionary
    ws1 = wb.active
    ws1.title = "Sheet Dictionary"
    write_sheet(ws1,
        ["name_pattern", "group", "role", "key_sheet", "analyze_by_default", "ignore_business_diff"],
        [[s.name_pattern, s.group, s.role, s.key_sheet, s.analyze_by_default, s.ignore_business_diff]
         for s in bd.sheet_dict]
    )

    # KPI Dictionary
    ws2 = wb.create_sheet("KPI Dictionary")
    write_sheet(ws2,
        ["name", "group", "level", "unit", "better_direction", "search_pattern", "search_in"],
        [[k.name, k.group, k.level, k.unit, k.better_direction, k.search_pattern, k.search_in]
         for k in bd.kpi_dict]
    )

    # Materiality Rules
    ws3 = wb.create_sheet("Materiality Rules")
    write_sheet(ws3,
        ["group", "abs_threshold", "pct_threshold"],
        [[grp, v.get("abs_threshold", ""), v.get("pct_threshold", "")]
         for grp, v in bd.materiality_rules.items()]
    )

    # Synonyms
    ws4 = wb.create_sheet("Synonyms")
    write_sheet(ws4,
        ["key", "synonyms (comma-separated)"],
        [[k, ", ".join(v)] for k, v in bd.synonyms.items()]
    )

    wb.save(str(path))
    log.info("Dictionary exported to Excel")


def import_from_excel(path: Path) -> tuple[BusinessDictionary | None, list[str]]:
    """Import dictionary from Excel file. Returns (dict, warnings)."""
    import openpyxl
    warnings: list[str] = []
    bd = load_dictionary()

    try:
        wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)

        if "Sheet Dictionary" in wb.sheetnames:
            ws = wb["Sheet Dictionary"]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            bd.sheet_dict = []
            for row in rows:
                if not row or row[0] is None:
                    continue
                try:
                    bd.sheet_dict.append(SheetEntry(
                        name_pattern=str(row[0]),
                        group=str(row[1]) if row[1] else "default",
                        role=str(row[2]) if row[2] else "other",
                        key_sheet=bool(row[3]) if len(row) > 3 else False,
                        analyze_by_default=bool(row[4]) if len(row) > 4 else True,
                        ignore_business_diff=bool(row[5]) if len(row) > 5 else False,
                    ))
                except Exception:
                    warnings.append("Ошибка в строке Sheet Dictionary")

        if "KPI Dictionary" in wb.sheetnames:
            ws = wb["KPI Dictionary"]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            bd.kpi_dict = []
            for row in rows:
                if not row or row[0] is None:
                    continue
                try:
                    bd.kpi_dict.append(KPIEntry(
                        name=str(row[0]),
                        group=str(row[1]) if row[1] else "Прочее",
                        level=int(row[2]) if row[2] else 2,
                        unit=str(row[3]) if row[3] else "",
                        better_direction=str(row[4]) if row[4] else "neutral",
                        search_pattern=str(row[5]) if row[5] else "",
                        search_in=str(row[6]) if len(row) > 6 and row[6] else "row_label",
                    ))
                except Exception:
                    warnings.append("Ошибка в строке KPI Dictionary")

        if "Materiality Rules" in wb.sheetnames:
            ws = wb["Materiality Rules"]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            bd.materiality_rules = {}
            for row in rows:
                if not row or row[0] is None:
                    continue
                try:
                    bd.materiality_rules[str(row[0])] = {
                        "abs_threshold": float(row[1]) if row[1] not in (None, "") else None,
                        "pct_threshold": float(row[2]) if row[2] not in (None, "") else None,
                    }
                except Exception:
                    warnings.append("Ошибка в строке Materiality Rules")

        wb.close()
        log.info(f"Dictionary imported from Excel: {len(warnings)} warnings")
        return bd, warnings
    except Exception as e:
        log.error(f"Dictionary import failed ({type(e).__name__})")
        return None, [f"Не удалось открыть файл: {type(e).__name__}"]


def get_sheet_group(bd: BusinessDictionary, sheet_name: str) -> str | None:
    for entry in bd.sheet_dict:
        if entry.name_pattern.lower() in sheet_name.lower():
            return entry.group
    return None


def is_key_sheet(bd: BusinessDictionary, sheet_name: str) -> bool:
    for entry in bd.sheet_dict:
        if entry.name_pattern.lower() in sheet_name.lower():
            return entry.key_sheet
    return False


def get_materiality(bd: BusinessDictionary, group: str) -> tuple[float | None, float | None]:
    """Returns (abs_threshold, pct_threshold) for a KPI group."""
    rule = bd.materiality_rules.get(group) or bd.materiality_rules.get("default", {})
    return rule.get("abs_threshold"), rule.get("pct_threshold")
