"""
KPI resolver utilities: parse cell address strings, export resolution table
to Excel for user editing, and import the edited table back.
"""
from __future__ import annotations
import re
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string

from fm_compare.core.models import CellAddress, KPIResolution
from fm_compare.security import safe_logger as log

_ADDR_RE = re.compile(
    r"^(?P<sheet>.+)!(?P<col>[A-Za-z]+)(?P<row>\d+)$"
)


def parse_cell_address(addr_str: str) -> CellAddress | None:
    """Parse 'Sheet!E42' into CellAddress. Returns None on bad input."""
    if not addr_str or "!" not in addr_str:
        return None
    m = _ADDR_RE.match(addr_str.strip())
    if not m:
        return None
    try:
        col = column_index_from_string(m.group("col"))
        row = int(m.group("row"))
        return CellAddress(sheet=m.group("sheet"), row=row, col=col)
    except Exception:
        return None


# ── Excel export ──────────────────────────────────────────────────────────────

_FILL_HEADER = PatternFill("solid", fgColor="4472C4")
_FILL_EVEN = PatternFill("solid", fgColor="EEF3FA")
_FONT_HEADER = Font(bold=True, color="FFFFFF", size=10)
_FONT_NORMAL = Font(size=10)
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

_COLUMNS = [
    ("kpi_name",  "KPI — Показатель",        200),
    ("kpi_group", "Группа",                   120),
    ("level",     "Ур.",                        40),
    ("label_v1",  "Найденная строка V1",       200),
    ("addr_v1",   "Ячейка V1 (редактируй)",   130),
    ("unit_v1",   "Ед. изм. V1 (редактируй)",  90),
    ("label_v2",  "Найденная строка V2",       200),
    ("addr_v2",   "Ячейка V2 (редактируй)",   130),
    ("unit_v2",   "Ед. изм. V2 (редактируй)",  90),
    ("source",    "Источник",                   80),
]


def export_resolutions_to_excel(
    resolutions: list[KPIResolution],
    output_path: Path,
) -> None:
    """Export KPI resolution table to Excel for user review/editing."""
    wb = Workbook()
    ws = wb.active
    ws.title = "KPI Resolution"

    # Instructions row
    ws.merge_cells("A1:H1")
    c = ws.cell(row=1, column=1,
                value="Проверьте и при необходимости исправьте адреса ячеек "
                      "(колонки 'Ячейка V1' и 'Ячейка V2'). "
                      "Формат: ИмяЛиста!E42. После правки сохраните файл и "
                      "загрузите его в программу через кнопку «Загрузить из Excel».")
    c.font = Font(size=10, italic=True, color="595959")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 36

    # Header row
    for ci, (_, header, _) in enumerate(_COLUMNS, 1):
        cell = ws.cell(row=2, column=ci, value=header)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = _ALIGN_CENTER
        cell.border = _THIN

    # Data rows
    for ri, res in enumerate(resolutions, 3):
        fill = _FILL_EVEN if ri % 2 == 0 else None
        vals = [
            res.kpi_name, res.kpi_group, res.kpi_level,
            res.label_v1, res.addr_v1, res.unit_v1,
            res.label_v2, res.addr_v2, res.unit_v2,
            res.source,
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.font = _FONT_NORMAL
            cell.alignment = _ALIGN_LEFT if ci in (1, 2, 4, 6) else _ALIGN_CENTER
            cell.border = _THIN
            if fill:
                cell.fill = fill

    # Column widths (approx pixel→char ratio 7)
    for ci, (_, _, px) in enumerate(_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = max(px / 7, 8)

    ws.freeze_panes = "A3"
    wb.save(str(output_path))
    log.info(f"Resolution table exported ({len(resolutions)} KPIs)")


# ── Excel import ──────────────────────────────────────────────────────────────

def import_resolutions_from_excel(
    path: Path,
) -> tuple[list[KPIResolution], list[str]]:
    """
    Import KPI resolution table from Excel.
    Returns (resolutions, errors). Errors are field-level validation messages.
    Cells that can't be parsed revert to empty string (not failed).
    """
    errors: list[str] = []
    resolutions: list[KPIResolution] = []

    try:
        wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    except Exception as e:
        return [], [f"Не удалось открыть файл: {e}"]

    if "KPI Resolution" not in wb.sheetnames:
        wb.close()
        return [], ["Лист 'KPI Resolution' не найден в файле."]

    ws = wb["KPI Resolution"]
    rows = list(ws.iter_rows(min_row=3, values_only=True))
    wb.close()

    for i, row in enumerate(rows, start=3):
        if not row or all(v is None for v in row):
            continue

        def _str(v: Any) -> str:
            return str(v).strip() if v is not None else ""

        kpi_name = _str(row[0])
        if not kpi_name:
            continue

        # Column layout (0-based): name, group, level,
        #   label_v1[3], addr_v1[4], unit_v1[5],
        #   label_v2[6], addr_v2[7], unit_v2[8], source[9]
        addr_v1_str = _str(row[4]) if len(row) > 4 else ""
        unit_v1_str = _str(row[5]) if len(row) > 5 else ""
        addr_v2_str = _str(row[7]) if len(row) > 7 else ""
        unit_v2_str = _str(row[8]) if len(row) > 8 else ""

        # Validate addresses if non-empty
        if addr_v1_str and parse_cell_address(addr_v1_str) is None:
            errors.append(
                f"Строка {i}: KPI «{kpi_name}» — неверный формат Ячейка V1: «{addr_v1_str}»"
            )
            addr_v1_str = ""

        if addr_v2_str and parse_cell_address(addr_v2_str) is None:
            errors.append(
                f"Строка {i}: KPI «{kpi_name}» — неверный формат Ячейка V2: «{addr_v2_str}»"
            )
            addr_v2_str = ""

        def _decompose(addr_s: str):
            if not addr_s:
                return "", None, None
            a = parse_cell_address(addr_s)
            if a is None:
                return "", None, None
            return a.sheet, a.row, a.col

        s1, r1, c1 = _decompose(addr_v1_str)
        s2, r2, c2 = _decompose(addr_v2_str)

        source = _str(row[9]) if len(row) > 9 else "manual"
        if source not in ("auto", "manual"):
            source = "manual"

        res = KPIResolution(
            kpi_name=kpi_name,
            kpi_group=_str(row[1]),
            kpi_level=int(row[2]) if row[2] is not None else 2,
            search_pattern="",
            label_v1=_str(row[3]) if len(row) > 3 else "",
            addr_v1=addr_v1_str,
            unit_v1=unit_v1_str,
            sheet_v1=s1, row_v1=r1, col_v1=c1,
            label_v2=_str(row[6]) if len(row) > 6 else "",
            addr_v2=addr_v2_str,
            unit_v2=unit_v2_str,
            sheet_v2=s2, row_v2=r2, col_v2=c2,
            source=source,
        )
        resolutions.append(res)

    log.info(f"Resolution table imported: {len(resolutions)} rows, {len(errors)} errors")
    return resolutions, errors


def resolutions_to_overrides(
    resolutions: list[KPIResolution],
    version: str,
) -> dict[str, CellAddress]:
    """
    Build an addr_overrides dict from a confirmed resolution list.
    version: "v1" or "v2"
    Only includes entries where address is fully specified.
    """
    overrides: dict[str, CellAddress] = {}
    for res in resolutions:
        if version == "v1":
            sheet, row, col = res.sheet_v1, res.row_v1, res.col_v1
        else:
            sheet, row, col = res.sheet_v2, res.row_v2, res.col_v2
        if sheet and row is not None and col is not None:
            overrides[res.kpi_name] = CellAddress(sheet=sheet, row=row, col=col)
    return overrides
