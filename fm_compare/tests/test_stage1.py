"""
Stage 1 tests: hierarchy, cross-check, upload endpoint, full workflow.

Unit tests use synthetic in-memory workbook data — no real files needed.
Upload / API tests use TestClient + minimal in-memory xlsx (openpyxl).
The smoke tests that use the real FM_Силикатный_*.xlsx skip when files absent.
"""
from __future__ import annotations

import io
import os
import pytest
from pathlib import Path

from fm_compare.core.excel_reader import SheetData, CellData, WorkbookData
from fm_compare.core.models import CellAddress
from fm_compare.core.hierarchy import build_hierarchy
from fm_compare.core.cross_check import cross_sheet_check


# ── XLSX factory ─────────────────────────────────────────────────────────────

def _make_xlsx(sheet_names: tuple[str, ...] = ("Sheet1",), with_data: bool = True) -> bytes:
    """Build a minimal valid .xlsx in memory using openpyxl."""
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    for name in sheet_names:
        ws = wb.create_sheet(name)
        if with_data:
            ws["A1"] = "Label"
            ws["B1"] = 42.0
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── TestClient fixture ────────────────────────────────────────────────────────

@pytest.fixture(scope="module")
def client():
    """FastAPI TestClient with auth gate disabled (APP_PASSWORD unset)."""
    import os
    os.environ.pop("APP_PASSWORD", None)   # open gate in test process
    from fastapi.testclient import TestClient
    from fm_compare.web.app import create_app
    with TestClient(create_app(), raise_server_exceptions=True) as c:
        yield c


# ── Helpers ──────────────────────────────────────────────────────────────────

def _cell(value=None, formula=None):
    return CellData(
        address=None, value=value, formula=formula,
        data_type=None, comment=None, is_hidden_row=False,
    )


def _sheet(cells: dict, name: str = "Sheet1") -> SheetData:
    rows = [r for r, _ in cells] or [1]
    cols = [c for _, c in cells] or [1]
    return SheetData(
        name=name, cells=cells, row_hidden={}, col_widths={},
        max_row=max(rows), max_col=max(cols), period_headers={},
    )


def _wb(sheets: dict) -> WorkbookData:
    return WorkbookData(info=None, sheets=sheets)


# ── build_hierarchy ───────────────────────────────────────────────────────────

class TestBuildHierarchy:

    def test_code_prefix_nesting(self):
        """5000 ⊃ 5400 ⊃ {5401, 5402} — three-level code-prefix hierarchy."""
        cells = {
            (1, 1): _cell("5000 Итого затраты"), (1, 3): _cell(-90e9),
            (2, 1): _cell("5400 Генподряд"),      (2, 3): _cell(-73e9),
            (3, 1): _cell("5401 Материалы"),       (3, 3): _cell(-30e9),
            (4, 1): _cell("5402 Работы"),           (4, 3): _cell(-43e9),
        }
        tree = build_hierarchy(_sheet(cells, "COST"))
        assert tree.sheet == "COST"
        assert len(tree.roots) == 1
        root = tree.roots[0]
        assert root.code == "5000"
        assert len(root.children) == 1
        c5400 = root.children[0]
        assert c5400.code == "5400"
        assert len(c5400.children) == 2
        child_codes = {ch.code for ch in c5400.children}
        assert child_codes == {"5401", "5402"}

    def test_by_code_index_populated(self):
        cells = {
            (1, 1): _cell("5000 Итого"), (1, 3): _cell(-100),
            (2, 1): _cell("5100 Часть"), (2, 3): _cell(-100),
        }
        tree = build_hierarchy(_sheet(cells))
        assert "5000" in tree.by_code
        assert "5100" in tree.by_code

    def test_empty_sheet_returns_empty_tree(self):
        tree = build_hierarchy(_sheet({}, "Empty"))
        assert tree.roots == []

    def test_underscore_code_parsed(self):
        """5400_СМР style code is recognised."""
        cells = {(1, 1): _cell("5400_СМР Генподряд"), (1, 3): _cell(-73e9)}
        tree = build_hierarchy(_sheet(cells))
        assert "5400" in tree.by_code


# ── cross_sheet_check ─────────────────────────────────────────────────────────

class TestCrossSheetCheck:

    def test_detects_discrepancy_between_sheets(self):
        """Same article on two sheets with ~10% difference → discrepancy."""
        cells_a = {(1, 1): _cell("5400 Генподряд"), (1, 3): _cell(-73e9)}
        cells_b = {(1, 1): _cell("5400 Генподряд"), (1, 3): _cell(-65e9)}
        wb = _wb({"COST": _sheet(cells_a, "COST"), "Себ20": _sheet(cells_b, "Себ20")})
        discrepancies = cross_sheet_check(wb, sheets=["COST", "Себ20"])
        assert len(discrepancies) >= 1
        d = discrepancies[0]
        assert d.kind == "cross_sheet"
        assert d.article == "5400"
        assert d.sheet_a in ("COST", "Себ20")
        assert d.sheet_b in ("COST", "Себ20")
        assert d.sheet_a != d.sheet_b

    def test_ignores_delta_within_rel_tolerance(self):
        """0.0001% difference is within default 1% rel tolerance → no discrepancy."""
        va = -73_000_000_000.0
        vb = va * (1 - 0.000001)
        cells_a = {(1, 1): _cell("5400 Генподряд"), (1, 3): _cell(va)}
        cells_b = {(1, 1): _cell("5400 Генподряд"), (1, 3): _cell(vb)}
        wb = _wb({"COST": _sheet(cells_a, "COST"), "Себ20": _sheet(cells_b, "Себ20")})
        assert cross_sheet_check(wb, sheets=["COST", "Себ20"]) == []

    def test_no_discrepancy_single_sheet(self):
        cells = {(1, 1): _cell("5400 Генподряд"), (1, 3): _cell(-73e9)}
        wb = _wb({"COST": _sheet(cells, "COST")})
        assert cross_sheet_check(wb, sheets=["COST"]) == []

    def test_worst_pair_reported_first(self):
        """When there are multiple discrepancies, largest delta comes first."""
        cells_a = {
            (1, 1): _cell("5400 Генподряд"), (1, 3): _cell(-73e9),
            (2, 1): _cell("5100 Доходы"),     (2, 3): _cell(100e9),
        }
        cells_b = {
            (1, 1): _cell("5400 Генподряд"), (1, 3): _cell(-50e9),   # Δ 23 млрд
            (2, 1): _cell("5100 Доходы"),     (2, 3): _cell(95e9),    # Δ 5 млрд
        }
        wb = _wb({"A": _sheet(cells_a, "A"), "B": _sheet(cells_b, "B")})
        discrepancies = cross_sheet_check(wb, sheets=["A", "B"])
        assert discrepancies[0].article == "5400"

    def test_returns_empty_on_empty_workbook(self):
        wb = _wb({})
        assert cross_sheet_check(wb) == []


# ── upload endpoint tests ─────────────────────────────────────────────────────

class TestUploadEndpoint:

    def test_upload_returns_job_id_and_sheet_list(self, client):
        """Two minimal xlsx files → 200 with job_id + non-empty sheet lists."""
        data = _make_xlsx(("RESUME", "COST"))
        resp = client.post(
            "/api/upload",
            files={"v1": ("model_v1.xlsx", data, "application/octet-stream"),
                   "v2": ("model_v2.xlsx", data, "application/octet-stream")},
        )
        assert resp.status_code == 200
        body = resp.json()
        assert "job_id" in body
        assert isinstance(body["job_id"], str) and body["job_id"]
        assert set(body["sheets_v1"]) == {"RESUME", "COST"}
        assert set(body["sheets_v2"]) == {"RESUME", "COST"}

    def test_upload_preserves_sheet_order(self, client):
        """Sheet names come back in workbook order."""
        sheets = ("RESUME", "COST", "CF", "PRICE")
        data = _make_xlsx(sheets)
        resp = client.post(
            "/api/upload",
            files={"v1": ("v1.xlsx", data, "application/octet-stream"),
                   "v2": ("v2.xlsx", data, "application/octet-stream")},
        )
        assert resp.status_code == 200
        body = resp.json()
        assert body["sheets_v1"] == list(sheets)

    def test_upload_xlsm_extension_accepted(self, client):
        """Macro-enabled .xlsm is allowed (same format as xlsx for read)."""
        data = _make_xlsx()
        resp = client.post(
            "/api/upload",
            files={"v1": ("model.xlsm", data, "application/octet-stream"),
                   "v2": ("model.xlsm", data, "application/octet-stream")},
        )
        assert resp.status_code == 200

    def test_upload_invalid_extension_rejected(self, client):
        """.csv files must be rejected with 422."""
        csv_bytes = b"label,value\nRevenue,100\n"
        resp = client.post(
            "/api/upload",
            files={"v1": ("model.csv", csv_bytes, "text/csv"),
                   "v2": ("model.csv", csv_bytes, "text/csv")},
        )
        assert resp.status_code == 422

    def test_upload_missing_v2_rejected(self, client):
        """Sending only v1 (no v2 field) → 422 from FastAPI validation."""
        data = _make_xlsx()
        resp = client.post(
            "/api/upload",
            files={"v1": ("model.xlsx", data, "application/octet-stream")},
        )
        assert resp.status_code == 422

    def test_upload_wrong_extension_only_v1(self, client):
        """.txt extension on v1 is rejected even if v2 is valid."""
        data = _make_xlsx()
        resp = client.post(
            "/api/upload",
            files={"v1": ("model.txt", b"garbage", "text/plain"),
                   "v2": ("model.xlsx", data, "application/octet-stream")},
        )
        assert resp.status_code == 422


# ── resolve-preview endpoint tests ───────────────────────────────────────────

class TestResolvePreview:

    def _upload(self, client) -> str:
        """Upload minimal files and return job_id."""
        data = _make_xlsx(("RESUME",))
        resp = client.post(
            "/api/upload",
            files={"v1": ("v1.xlsx", data, "application/octet-stream"),
                   "v2": ("v2.xlsx", data, "application/octet-stream")},
        )
        assert resp.status_code == 200
        return resp.json()["job_id"]

    def test_resolve_preview_returns_resolutions(self, client):
        """resolve-preview on a minimal file returns a list (possibly empty)."""
        job_id = self._upload(client)
        resp = client.post(
            f"/api/{job_id}/resolve-preview",
            json={"sheets_v1": ["RESUME"], "sheets_v2": ["RESUME"]},
        )
        assert resp.status_code == 200
        body = resp.json()
        assert "resolutions" in body
        assert isinstance(body["resolutions"], list)
        assert "corrections" in body  # LLM validator response (empty when gateway down)

    def test_resolve_preview_unknown_job_returns_404(self, client):
        resp = client.post(
            "/api/nonexistent-job/resolve-preview",
            json={"sheets_v1": ["Sheet1"], "sheets_v2": ["Sheet1"]},
        )
        assert resp.status_code == 404

    def test_resolve_preview_empty_sheets_returns_422(self, client):
        job_id = self._upload(client)
        resp = client.post(
            f"/api/{job_id}/resolve-preview",
            json={"sheets_v1": [], "sheets_v2": ["RESUME"]},
        )
        assert resp.status_code == 422


# ── healthcheck ───────────────────────────────────────────────────────────────

def test_healthz(client):
    assert client.get("/healthz").status_code == 200


# ── real-file smoke test (skips when xlsx not present) ───────────────────────

_DATA_DIR = Path(__file__).parent.parent   # fm_compare/
_FM_V1 = _DATA_DIR / "FM_Силикатный_1Q2026.xlsx"
_FM_V2 = _DATA_DIR / "FM_Силикатный_4Q2025.xlsx"


@pytest.mark.skipif(
    not (_FM_V1.exists() and _FM_V2.exists()),
    reason="Real FM files not found (run locally with fm_compare/FM_Силикатный_*.xlsx)",
)
def test_build_dashboard_real_files():
    """End-to-end smoke: build_dashboard returns kpis + dates + discrepancies."""
    from fm_compare.core.business_dictionary import load_dictionary
    from fm_compare.web.dashboard import build_dashboard

    bd = load_dictionary()
    sheets = ["RESUME"]
    payload = build_dashboard(_FM_V1, _FM_V2, bd, sheets, sheets)

    assert "kpis" in payload
    assert "date_v1" in payload
    assert "date_v2" in payload
    assert "discrepancies" in payload
    assert isinstance(payload["kpis"], list)
    assert len(payload["kpis"]) > 0
    for kpi in payload["kpis"]:
        assert "kpi_name" in kpi
        assert "value_v1" in kpi
        assert "value_v2" in kpi


@pytest.mark.skipif(
    not (_FM_V1.exists() and _FM_V2.exists()),
    reason="Real FM files not found",
)
def test_upload_and_dashboard_real_files(client):
    """Full HTTP workflow with real FM files: upload → resolve-preview → dashboard."""
    # Step 1: upload
    with open(_FM_V1, "rb") as f1, open(_FM_V2, "rb") as f2:
        resp = client.post(
            "/api/upload",
            files={"v1": (_FM_V1.name, f1.read(), "application/octet-stream"),
                   "v2": (_FM_V2.name, f2.read(), "application/octet-stream")},
        )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    job_id = body["job_id"]
    sheets_v1 = body["sheets_v1"]
    sheets_v2 = body["sheets_v2"]
    assert len(sheets_v1) > 10, "FM file should have many sheets"
    assert "RESUME" in sheets_v1

    # Step 2: resolve-preview (only RESUME sheet for speed)
    resp = client.post(
        f"/api/{job_id}/resolve-preview",
        json={"sheets_v1": ["RESUME"], "sheets_v2": ["RESUME"]},
    )
    assert resp.status_code == 200, resp.text
    preview = resp.json()
    assert "resolutions" in preview
    assert len(preview["resolutions"]) > 0, "Should find at least one KPI on RESUME"
    assert "corrections" in preview   # empty list when gateway unavailable

    # Step 3: dashboard
    resp = client.get(f"/api/{job_id}/dashboard")
    assert resp.status_code == 200, resp.text
    dash = resp.json()
    assert "kpis" in dash
    assert "discrepancies" in dash
    assert "date_v1" in dash
    assert "date_v2" in dash
    assert len(dash["kpis"]) > 0
    # Every KPI row must have the required fields
    required = {"kpi_name", "kpi_group", "value_v1", "value_v2", "delta", "delta_pct"}
    for kpi in dash["kpis"]:
        assert required.issubset(kpi.keys()), f"Missing keys in {kpi['kpi_name']}"
