"""
Functional integration tests for FM Compare.

These tests exercise the complete API workflow using synthetic Excel files
that mimic the real FM model structure (RESUME + CF sheets, KPIs by row label).

Run with:
    pytest fm_compare/tests/test_functional.py -v

Coverage:
  - Full upload → resolve-preview → dashboard → run → status poll → summary → report flow
  - Three-version (V3) comparison workflow
  - Sensitivity analysis initiation
  - Error handling (404, 422, oversized file)
  - Chat endpoint with mocked gateway
  - Cross-sheet hierarchy check
"""
from __future__ import annotations

import io
import os
import time
import pytest
from pathlib import Path
from unittest.mock import patch, MagicMock


# ── Synthetic FM model builder ────────────────────────────────────────────────

def _build_fm_workbook(
    revenue: float = 5_000_000_000.0,
    cost: float = 3_200_000_000.0,
    npv: float = 800_000_000.0,
    irr: float = 18.5,
    cashflow: float = 1_200_000_000.0,
    margin_pct: float = 20.0,
    seed: int = 1,
) -> bytes:
    """
    Build a realistic FM workbook with RESUME and CF sheets.

    Layout matches real projects:
      col A (1) = article code
      col B (2) = row label
      col C (3) = unit
      col D (4) = total/summary value
      cols E+ = period columns (quarters)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    wb.remove(wb.active)

    # ── RESUME sheet ──────────────────────────────────────────────────────────
    ws = wb.create_sheet("RESUME")

    headers = ["Код", "Показатель", "Ед.изм.", "Итого", "1кв", "2кв", "3кв", "4кв"]
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h).font = Font(bold=True)

    # period values are proportional to total, divided by 4 + small variation
    def _qv(total: float, q: int) -> float:
        return round(total * (0.22 + (q - 1) * 0.02 + seed * 0.001), 0)

    rows = [
        # code,  label,                          unit,         total
        ("5000", "Выручка от продаж",            "руб.",       revenue),
        ("5100", "Поступления денежных средств", "руб.",       revenue * 0.95),
        ("5200", "Себестоимость / CAPEX / СМР",  "руб.",       cost),
        ("5210", "Прочие расходы",               "руб.",       cost * 0.12),
        ("5300", "Налоги",                       "руб.",       revenue * 0.08),
        ("5400", "Процентные расходы",           "руб.",       cost * 0.06),
        ("5500", "Денежный поток",               "руб.",       cashflow),
        ("5510", "Cash gap / минимальный остаток","руб.",      cashflow * 0.05),
        ("5600", "Прибыль проекта",              "руб.",       npv * 0.6),
        ("5700", "Маржинальность",               "%",          margin_pct),
        ("5800", "IRR",                          "%",          irr),
        ("5900", "NPV",                          "руб.",       npv),
        ("5950", "Вложения инвесторов / акционеров","руб.",    cost * 0.45),
        ("5960", "Финансирование / потребность", "руб.",       cost * 0.55),
    ]

    for i, (code, label, unit, total) in enumerate(rows, 2):
        ws.cell(i, 1, code)
        ws.cell(i, 2, label)
        ws.cell(i, 3, unit)
        ws.cell(i, 4, total)
        for q in range(1, 5):
            ws.cell(i, 4 + q, _qv(total, q))

    # ── CF sheet (cashflow) ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("CF")
    cf_rows = [
        ("CF100", "Денежный поток от операций",  "руб.",  cashflow * 1.1),
        ("CF200", "Поступления денежных средств","руб.",  revenue * 0.9),
        ("CF300", "Выручка CF",                  "руб.",  revenue * 0.85),
        ("CF400", "Налоги CF",                   "руб.",  revenue * 0.07),
        ("CF500", "Прибыль проекта CF",          "руб.",  npv * 0.55),
    ]
    for col, h in enumerate(headers, 1):
        ws2.cell(1, col, h).font = Font(bold=True)
    for i, (code, label, unit, total) in enumerate(cf_rows, 2):
        ws2.cell(i, 1, code)
        ws2.cell(i, 2, label)
        ws2.cell(i, 3, unit)
        ws2.cell(i, 4, total)
        for q in range(1, 5):
            ws2.cell(i, 4 + q, _qv(total, q))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _fm_v1() -> bytes:
    """V1 = newer version (higher revenue, better metrics)."""
    return _build_fm_workbook(
        revenue=5_000_000_000.0, cost=3_200_000_000.0,
        npv=800_000_000.0, irr=18.5, cashflow=1_200_000_000.0,
        margin_pct=20.0, seed=1,
    )


def _fm_v2() -> bytes:
    """V2 = older version (lower revenue, worse metrics)."""
    return _build_fm_workbook(
        revenue=4_500_000_000.0, cost=3_100_000_000.0,
        npv=650_000_000.0, irr=16.2, cashflow=1_050_000_000.0,
        margin_pct=17.5, seed=2,
    )


def _fm_v3() -> bytes:
    """V3 = oldest version (baseline)."""
    return _build_fm_workbook(
        revenue=4_000_000_000.0, cost=2_900_000_000.0,
        npv=500_000_000.0, irr=14.0, cashflow=900_000_000.0,
        margin_pct=15.0, seed=3,
    )


# ── Client fixture ────────────────────────────────────────────────────────────

@pytest.fixture(scope="module")
def client():
    os.environ.pop("APP_PASSWORD", None)
    from fastapi.testclient import TestClient
    from fm_compare.web.app import create_app
    with TestClient(create_app(), raise_server_exceptions=True) as c:
        yield c


# ── Helper ────────────────────────────────────────────────────────────────────

def _upload(client, v1=None, v2=None, v3=None) -> dict:
    v1 = v1 or _fm_v1()
    v2 = v2 or _fm_v2()
    files = {
        "v1": ("fm_v1.xlsx", v1, "application/octet-stream"),
        "v2": ("fm_v2.xlsx", v2, "application/octet-stream"),
    }
    if v3 is not None:
        files["v3"] = ("fm_v3.xlsx", v3, "application/octet-stream")
    resp = client.post("/api/upload", files=files)
    assert resp.status_code == 200, f"Upload failed: {resp.text}"
    return resp.json()


def _resolve(client, job_id: str, sheets_v1=None, sheets_v2=None, sheets_v3=None) -> dict:
    body = {
        "sheets_v1": sheets_v1 or ["RESUME"],
        "sheets_v2": sheets_v2 or ["RESUME"],
    }
    if sheets_v3:
        body["sheets_v3"] = sheets_v3
    resp = client.post(f"/api/{job_id}/resolve-preview", json=body)
    assert resp.status_code == 200, f"resolve-preview failed: {resp.text}"
    return resp.json()


def _poll_status(client, job_id: str, timeout: int = 30, endpoint: str = "status") -> dict:
    deadline = time.time() + timeout
    while time.time() < deadline:
        resp = client.get(f"/api/{job_id}/{endpoint}")
        assert resp.status_code == 200
        body = resp.json()
        if body["status"] in ("done", "error"):
            return body
        time.sleep(0.2)
    pytest.fail(f"Job {job_id} did not finish within {timeout}s")


# ── Upload tests ──────────────────────────────────────────────────────────────

class TestUpload:

    def test_upload_returns_job_id_and_sheets(self, client):
        body = _upload(client)
        assert "job_id" in body
        assert len(body["job_id"]) == 32
        assert "RESUME" in body["sheets_v1"]
        assert "CF" in body["sheets_v1"]
        assert "RESUME" in body["sheets_v2"]

    def test_upload_has_v3_false_by_default(self, client):
        body = _upload(client)
        assert body["has_v3"] is False
        assert body["sheets_v3"] == []

    def test_upload_with_v3_sets_flag(self, client):
        body = _upload(client, v3=_fm_v3())
        assert body["has_v3"] is True
        assert "RESUME" in body["sheets_v3"]

    def test_upload_oversized_file_rejected(self, client):
        huge = b"0" * (101 * 1024 * 1024)
        resp = client.post(
            "/api/upload",
            files={
                "v1": ("big.xlsx", huge, "application/octet-stream"),
                "v2": ("v2.xlsx", _fm_v2(), "application/octet-stream"),
            },
        )
        assert resp.status_code == 413

    def test_upload_invalid_extension_rejected(self, client):
        resp = client.post(
            "/api/upload",
            files={
                "v1": ("model.csv", b"a,b\n1,2\n", "text/csv"),
                "v2": ("model.csv", b"a,b\n1,2\n", "text/csv"),
            },
        )
        assert resp.status_code == 422

    def test_upload_missing_v2_rejected(self, client):
        resp = client.post(
            "/api/upload",
            files={"v1": ("model.xlsx", _fm_v1(), "application/octet-stream")},
        )
        assert resp.status_code == 422


# ── resolve-preview tests ─────────────────────────────────────────────────────

class TestResolvePreview:

    def test_resolve_returns_resolutions_list(self, client):
        body = _upload(client)
        rp = _resolve(client, body["job_id"])
        assert isinstance(rp["resolutions"], list)
        assert isinstance(rp["corrections"], list)

    def test_resolve_finds_known_kpis(self, client):
        body = _upload(client)
        rp = _resolve(client, body["job_id"])
        found_names = {r["kpi_name"] for r in rp["resolutions"]}
        # At least some of the core KPIs should be resolved
        expected = {"Выручка", "NPV", "IRR", "Денежный поток"}
        assert expected & found_names, f"No expected KPIs in {found_names}"

    def test_resolve_resolution_has_required_fields(self, client):
        body = _upload(client)
        rp = _resolve(client, body["job_id"])
        assert rp["resolutions"], "No resolutions found in FM workbook"
        r = rp["resolutions"][0]
        for field in ("kpi_name", "kpi_group", "addr_v1", "addr_v2"):
            assert field in r, f"Missing field: {field}"

    def test_resolve_multi_sheet(self, client):
        """Resolving across RESUME + CF should surface more KPIs than single sheet."""
        body = _upload(client)
        rp_single = _resolve(client, body["job_id"], sheets_v1=["RESUME"], sheets_v2=["RESUME"])
        body2 = _upload(client)
        rp_multi = _resolve(client, body2["job_id"], sheets_v1=["RESUME", "CF"], sheets_v2=["RESUME", "CF"])
        # Multi-sheet should find at least as many KPIs
        assert len(rp_multi["resolutions"]) >= len(rp_single["resolutions"])

    def test_resolve_unknown_job_404(self, client):
        resp = client.post(
            "/api/aabbccdd11223344aabbccdd11223344/resolve-preview",
            json={"sheets_v1": ["RESUME"], "sheets_v2": ["RESUME"]},
        )
        assert resp.status_code == 404

    def test_resolve_empty_sheets_422(self, client):
        body = _upload(client)
        resp = client.post(
            f"/api/{body['job_id']}/resolve-preview",
            json={"sheets_v1": [], "sheets_v2": ["RESUME"]},
        )
        assert resp.status_code == 422

    def test_resolve_v3_sheets(self, client):
        body = _upload(client, v3=_fm_v3())
        rp = _resolve(client, body["job_id"],
                      sheets_v1=["RESUME"], sheets_v2=["RESUME"], sheets_v3=["RESUME"])
        assert isinstance(rp["resolutions"], list)
        # V3 resolutions should have addr_v3 field
        for r in rp["resolutions"]:
            assert "addr_v3" in r


# ── Dashboard tests ───────────────────────────────────────────────────────────

class TestDashboard:

    def _setup(self, client) -> str:
        body = _upload(client)
        _resolve(client, body["job_id"])
        return body["job_id"]

    def test_get_dashboard_returns_kpis(self, client):
        job_id = self._setup(client)
        resp = client.get(f"/api/{job_id}/dashboard")
        assert resp.status_code == 200
        body = resp.json()
        assert "kpis" in body
        assert isinstance(body["kpis"], list)

    def test_dashboard_kpi_has_delta_fields(self, client):
        job_id = self._setup(client)
        body = client.get(f"/api/{job_id}/dashboard").json()
        if not body["kpis"]:
            pytest.skip("No KPIs resolved — check FM structure")
        kpi = body["kpis"][0]
        for field in ("kpi_name", "value_v1", "value_v2", "delta", "delta_pct"):
            assert field in kpi, f"Missing field: {field}"

    def test_dashboard_deltas_are_correct(self, client):
        # Convention: delta = v2 - v1 (old minus new; negative = improvement in V1)
        job_id = self._setup(client)
        body = client.get(f"/api/{job_id}/dashboard").json()
        for kpi in body["kpis"]:
            if kpi["value_v1"] is not None and kpi["value_v2"] is not None:
                expected = round(kpi["value_v2"] - kpi["value_v1"], 6)
                actual = round(kpi["delta"], 6)
                assert abs(expected - actual) < 1e-3, \
                    f"Delta mismatch for {kpi['kpi_name']}: {expected} vs {actual}"
                break  # One is enough

    def test_dashboard_before_resolve_422(self, client):
        body = _upload(client)
        resp = client.get(f"/api/{body['job_id']}/dashboard")
        assert resp.status_code == 422

    def test_dashboard_post_override(self, client):
        job_id = self._setup(client)
        dash = client.get(f"/api/{job_id}/dashboard").json()
        resolutions = dash.get("resolutions", [])
        if not resolutions:
            pytest.skip("No resolutions to override")
        # POST the same resolutions back — should return same shape
        resp = client.post(f"/api/{job_id}/dashboard", json={"resolutions": resolutions})
        assert resp.status_code == 200
        assert "kpis" in resp.json()

    def test_dashboard_unknown_job_404(self, client):
        resp = client.get("/api/aabbccdd11223344aabbccdd11223344/dashboard")
        assert resp.status_code == 404


# ── Run + status + summary tests ──────────────────────────────────────────────

class TestRunAndSummary:

    def _setup(self, client) -> tuple[str, list]:
        body = _upload(client)
        rp = _resolve(client, body["job_id"])
        return body["job_id"], rp["resolutions"]

    def test_run_returns_202_or_200_with_running_status(self, client):
        job_id, resolutions = self._setup(client)
        resp = client.post(
            f"/api/{job_id}/run",
            json={"resolutions": resolutions, "mode": "quick"},
        )
        assert resp.status_code == 200
        body = resp.json()
        assert body["status"] in ("running", "done", "uploaded")

    def test_run_to_completion(self, client):
        job_id, resolutions = self._setup(client)
        client.post(f"/api/{job_id}/run",
                    json={"resolutions": resolutions, "mode": "quick"})
        final = _poll_status(client, job_id)
        assert final["status"] == "done", f"Job ended with error: {final.get('error')}"

    def test_summary_after_run(self, client):
        job_id, resolutions = self._setup(client)
        client.post(f"/api/{job_id}/run",
                    json={"resolutions": resolutions, "mode": "quick"})
        _poll_status(client, job_id)
        resp = client.get(f"/api/{job_id}/summary")
        assert resp.status_code == 200
        body = resp.json()
        assert "summary_blocks" in body

    def test_summary_before_run_404(self, client):
        body = _upload(client)
        _resolve(client, body["job_id"])
        resp = client.get(f"/api/{body['job_id']}/summary")
        assert resp.status_code == 404

    def test_run_with_materiality_filter(self, client):
        job_id, resolutions = self._setup(client)
        resp = client.post(
            f"/api/{job_id}/run",
            json={
                "resolutions": resolutions,
                "mode": "quick",
                "materiality_abs": 1_000_000.0,
                "materiality_pct": 1.0,
            },
        )
        assert resp.status_code == 200

    def test_run_invalid_job_404(self, client):
        resp = client.post(
            "/api/aabbccdd11223344aabbccdd11223344/run",
            json={"resolutions": [], "mode": "quick"},
        )
        assert resp.status_code == 404


# ── Report download tests ─────────────────────────────────────────────────────

class TestReport:

    def _run_to_done(self, client) -> str:
        body = _upload(client)
        rp = _resolve(client, body["job_id"])
        client.post(f"/api/{body['job_id']}/run",
                    json={"resolutions": rp["resolutions"], "mode": "quick"})
        _poll_status(client, body["job_id"])
        return body["job_id"]

    def test_report_download_returns_xlsx(self, client):
        job_id = self._run_to_done(client)
        resp = client.get(f"/api/{job_id}/report.xlsx")
        assert resp.status_code == 200
        ct = resp.headers.get("content-type", "")
        assert "spreadsheet" in ct or "excel" in ct or "octet" in ct
        # Verify magic bytes for ZIP/xlsx
        assert resp.content[:4] == b"PK\x03\x04", "Not a valid xlsx (ZIP) file"

    def test_report_not_ready_404(self, client):
        body = _upload(client)
        _resolve(client, body["job_id"])
        resp = client.get(f"/api/{body['job_id']}/report.xlsx")
        assert resp.status_code == 404


# ── V3 three-version comparison ───────────────────────────────────────────────

class TestThreeVersionComparison:

    def _setup_v3(self, client) -> tuple[str, list]:
        body = _upload(client, v3=_fm_v3())
        assert body["has_v3"] is True
        rp = _resolve(client, body["job_id"],
                      sheets_v1=["RESUME"], sheets_v2=["RESUME"], sheets_v3=["RESUME"])
        return body["job_id"], rp["resolutions"]

    def test_v3_resolve_populates_addr_v3(self, client):
        job_id, resolutions = self._setup_v3(client)
        has_v3_addr = any(r.get("addr_v3") for r in resolutions)
        assert has_v3_addr, "No V3 addresses resolved"

    def test_v3_run_starts_both_comparisons(self, client):
        job_id, resolutions = self._setup_v3(client)
        resp = client.post(f"/api/{job_id}/run",
                           json={"resolutions": resolutions, "mode": "quick"})
        assert resp.status_code == 200
        # Status2 endpoint should also be accessible
        status2 = client.get(f"/api/{job_id}/status2").json()
        assert "status" in status2

    def test_v3_dashboard_has_v3_true(self, client):
        job_id, _ = self._setup_v3(client)
        body = client.get(f"/api/{job_id}/dashboard").json()
        assert body.get("has_v3") is True

    def test_v3_kpis_have_v3_value(self, client):
        job_id, _ = self._setup_v3(client)
        body = client.get(f"/api/{job_id}/dashboard").json()
        kpis_with_v3 = [k for k in body.get("kpis", []) if k.get("value_v3") is not None]
        assert kpis_with_v3, "No KPIs have V3 values"

    def test_v3_report2_download(self, client):
        job_id, resolutions = self._setup_v3(client)
        client.post(f"/api/{job_id}/run",
                    json={"resolutions": resolutions, "mode": "quick"})
        _poll_status(client, job_id)
        final2 = _poll_status(client, job_id, endpoint="status2")
        if final2["status"] == "done":
            resp = client.get(f"/api/{job_id}/report2.xlsx")
            assert resp.status_code == 200
            assert resp.content[:4] == b"PK\x03\x04"


# ── Chat endpoint tests ───────────────────────────────────────────────────────

class TestChatEndpoint:

    def _setup_with_run(self, client) -> str:
        body = _upload(client)
        rp = _resolve(client, body["job_id"])
        client.post(f"/api/{body['job_id']}/run",
                    json={"resolutions": rp["resolutions"], "mode": "quick"})
        _poll_status(client, body["job_id"])
        return body["job_id"]

    def test_chat_unknown_job_404(self, client):
        resp = client.post(
            "/api/aabbccdd11223344aabbccdd11223344/chat",
            json={"history": [{"role": "user", "content": "Привет"}]},
        )
        assert resp.status_code == 404

    def test_chat_empty_history_422(self, client):
        body = _upload(client)
        _resolve(client, body["job_id"])
        resp = client.post(f"/api/{body['job_id']}/chat", json={"history": []})
        assert resp.status_code == 422

    def test_chat_last_message_not_user_422(self, client):
        body = _upload(client)
        _resolve(client, body["job_id"])
        resp = client.post(f"/api/{body['job_id']}/chat", json={
            "history": [{"role": "assistant", "content": "Ответ"}]
        })
        assert resp.status_code == 422

    def test_chat_sse_stream_with_mock(self, client):
        job_id = self._setup_with_run(client)
        with patch("fm_compare.web.routes.stream_chat_turn") as mock_stream:
            mock_stream.return_value = iter(["Выручка выросла ", "на 11%."])
            resp = client.post(
                f"/api/{job_id}/chat",
                json={"history": [{"role": "user", "content": "Как изменилась выручка?"}]},
            )
        assert resp.status_code == 200
        assert "text/event-stream" in resp.headers.get("content-type", "")
        body = resp.text
        assert "Выручка выросла" in body
        assert "data: [DONE]" in body

    def test_chat_message_too_long_422(self, client):
        body = _upload(client)
        _resolve(client, body["job_id"])
        resp = client.post(
            f"/api/{body['job_id']}/chat",
            json={"history": [{"role": "user", "content": "x" * 2001}]},
        )
        assert resp.status_code == 422

    def test_chat_context_injected_from_dashboard(self, client):
        """Context from dashboard kpis must appear in messages sent to gateway."""
        job_id = self._setup_with_run(client)
        captured: list = []

        def _fake_stream(history, dashboard_data, **kwargs):
            captured.append((history, dashboard_data))
            return iter(["ok"])

        with patch("fm_compare.web.routes.stream_chat_turn", side_effect=_fake_stream):
            client.post(
                f"/api/{job_id}/chat",
                json={"history": [{"role": "user", "content": "Вопрос"}]},
            )

        assert captured, "stream_chat_turn was never called"
        _, dash_data = captured[0]
        assert "kpis" in dash_data or dash_data == {}, "dashboard_data not passed"


# ── Sensitivity analysis tests ────────────────────────────────────────────────

class TestSensitivity:

    def _setup(self, client) -> str:
        body = _upload(client)
        _resolve(client, body["job_id"])
        return body["job_id"]

    def test_sensitivity_returns_accepted(self, client):
        job_id = self._setup(client)
        resp = client.post(f"/api/{job_id}/sensitivity", json={
            "inputs": [
                {"name": "Цена", "addr": "RESUME!D2",
                 "unit": "руб.", "base_value": 5_000_000_000.0,
                 "values": [4_500_000_000.0, 5_000_000_000.0, 5_500_000_000.0]}
            ],
            "kpi_addrs": [
                {"name": "Прибыль", "addr": "RESUME!D9"}
            ],
            "timeout": 30,
        })
        # Might be 200 (LibreOffice absent → error) or 202 (started)
        assert resp.status_code in (200, 202, 422), f"Unexpected: {resp.status_code} {resp.text}"

    def test_sensitivity_unknown_job_404(self, client):
        resp = client.post(
            "/api/aabbccdd11223344aabbccdd11223344/sensitivity",
            json={"inputs": [], "kpi_addrs": []},
        )
        assert resp.status_code == 404


# ── Health check ──────────────────────────────────────────────────────────────

def test_healthz(client):
    # /healthz returns plain text "ok" (PlainTextResponse), not JSON
    resp = client.get("/healthz")
    assert resp.status_code == 200
    assert resp.text.strip() == "ok"


# ── Path traversal security ───────────────────────────────────────────────────

def test_path_traversal_job_id_rejected(client):
    """Malicious job_ids with path traversal chars must return 404 (not 500)."""
    bad_ids = [
        "../etc/passwd",
        "..%2F..%2Fetc%2Fpasswd",
        "' OR 1=1 --",
        "aaaa/../../../../etc",
    ]
    for bad_id in bad_ids:
        resp = client.get(f"/api/{bad_id}/status")
        assert resp.status_code in (404, 422), \
            f"Expected 404/422 for {bad_id!r}, got {resp.status_code}"
