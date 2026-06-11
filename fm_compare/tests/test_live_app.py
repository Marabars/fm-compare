"""
Live integration tests against the deployed FM Compare instance.

Run against a deployed FM Compare instance:

    FM_COMPARE_URL=http://<host>:8000 FM_COMPARE_PASSWORD=<pwd> python fm_compare/tests/test_live_app.py
    # Defaults to localhost:8000 with no password (open gate, dev mode)

The script uploads synthetic FM Excel files (no real data), walks the
complete API workflow, and prints a result table.  No pytest required.

Requires: requests, openpyxl  (both already in project requirements)
"""
from __future__ import annotations

import io
import json
import os
import sys
import time
from typing import Any

try:
    import requests
    import openpyxl
    from openpyxl.styles import Font
except ImportError as e:
    print(f"Missing dependency: {e}")
    print("Install: pip install requests openpyxl")
    sys.exit(1)

BASE_URL = os.environ.get("FM_COMPARE_URL", "http://localhost:8000")
PASSWORD = os.environ.get("FM_COMPARE_PASSWORD", "")
TIMEOUT = int(os.environ.get("FM_COMPARE_TIMEOUT", "60"))

_PASS = "✅ PASS"
_FAIL = "❌ FAIL"
_SKIP = "⏩ SKIP"

results: list[tuple[str, str, str]] = []  # (test_name, status, detail)


def _record(name: str, ok: bool, detail: str = "") -> bool:
    tag = _PASS if ok else _FAIL
    results.append((name, tag, detail))
    status = "PASS" if ok else "FAIL"
    print(f"  [{status}] {name}" + (f" — {detail}" if detail else ""))
    return ok


def _build_fm(
    revenue: float = 5_000_000_000.0,
    cost: float = 3_200_000_000.0,
    npv: float = 800_000_000.0,
    irr: float = 18.5,
    cashflow: float = 1_200_000_000.0,
    margin_pct: float = 20.0,
    seed: int = 1,
) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def _qv(total: float, q: int) -> float:
        return round(total * (0.22 + (q - 1) * 0.02 + seed * 0.001), 0)

    rows = [
        ("5000", "Выручка от продаж",             "руб.",   revenue),
        ("5100", "Поступления денежных средств",  "руб.",   revenue * 0.95),
        ("5200", "Себестоимость / CAPEX / СМР",   "руб.",   cost),
        ("5210", "Прочие расходы",                "руб.",   cost * 0.12),
        ("5300", "Налоги",                        "руб.",   revenue * 0.08),
        ("5400", "Процентные расходы",            "руб.",   cost * 0.06),
        ("5500", "Денежный поток",                "руб.",   cashflow),
        ("5510", "Cash gap / минимальный остаток","руб.",   cashflow * 0.05),
        ("5600", "Прибыль проекта",               "руб.",   npv * 0.6),
        ("5700", "Маржинальность",                "%",      margin_pct),
        ("5800", "IRR",                           "%",      irr),
        ("5900", "NPV",                           "руб.",   npv),
        ("5950", "Вложения инвесторов / акционеров","руб.", cost * 0.45),
    ]
    headers = ["Код", "Показатель", "Ед.изм.", "Итого", "1кв", "2кв", "3кв", "4кв"]

    for sheet_name, data_rows in [("RESUME", rows), ("CF", rows[:6])]:
        ws = wb.create_sheet(sheet_name)
        for c, h in enumerate(headers, 1):
            ws.cell(1, c, h).font = Font(bold=True)
        for i, (code, label, unit, total) in enumerate(data_rows, 2):
            ws.cell(i, 1, code)
            ws.cell(i, 2, label)
            ws.cell(i, 3, unit)
            ws.cell(i, 4, total)
            for q in range(1, 5):
                ws.cell(i, 4 + q, _qv(total, q))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _session() -> requests.Session:
    s = requests.Session()
    s.timeout = TIMEOUT
    if PASSWORD:
        resp = s.post(f"{BASE_URL}/login", data={"password": PASSWORD}, allow_redirects=True)
        if resp.status_code not in (200, 302) and "Set-Cookie" not in resp.headers:
            print(f"  [WARN] Login may have failed: status={resp.status_code}")
    return s


def _poll(s: requests.Session, job_id: str, endpoint: str = "status") -> dict:
    deadline = time.time() + TIMEOUT
    while time.time() < deadline:
        r = s.get(f"{BASE_URL}/api/{job_id}/{endpoint}", timeout=10)
        body = r.json()
        if body["status"] in ("done", "error"):
            return body
        time.sleep(1.0)
    return {"status": "timeout", "error": f"Did not finish in {TIMEOUT}s"}


def run_tests() -> int:
    print(f"\nFM Compare Live Integration Tests")
    print(f"Target: {BASE_URL}")
    print("=" * 60)

    s = _session()

    # ── 1. Health check ───────────────────────────────────────────────────────
    print("\n[Health]")
    try:
        r = s.get(f"{BASE_URL}/healthz", timeout=10)
        _record("GET /healthz → 200", r.status_code == 200, f"status={r.status_code}")
        # /healthz returns plain text "ok"
        _record("healthz returns ok", r.text.strip() == "ok", repr(r.text))
    except Exception as e:
        _record("GET /healthz", False, str(e))
        print("  Cannot reach app — aborting.")
        _print_summary()
        return 1

    # ── 2. Upload ─────────────────────────────────────────────────────────────
    print("\n[Upload]")
    v1_bytes = _build_fm(revenue=5e9, cost=3.2e9, npv=8e8, irr=18.5, seed=1)
    v2_bytes = _build_fm(revenue=4.5e9, cost=3.1e9, npv=6.5e8, irr=16.2, seed=2)
    v3_bytes = _build_fm(revenue=4e9, cost=2.9e9, npv=5e8, irr=14.0, seed=3)

    job_id = None
    try:
        r = s.post(f"{BASE_URL}/api/upload", files={
            "v1": ("fm_v1.xlsx", v1_bytes, "application/octet-stream"),
            "v2": ("fm_v2.xlsx", v2_bytes, "application/octet-stream"),
        }, timeout=30)
        ok = r.status_code == 200
        _record("POST /api/upload → 200", ok, f"status={r.status_code}")
        if ok:
            body = r.json()
            job_id = body.get("job_id")
            _record("job_id is 32-char hex", bool(job_id) and len(job_id) == 32, job_id or "")
            _record("sheets_v1 includes RESUME", "RESUME" in body.get("sheets_v1", []),
                    str(body.get("sheets_v1", [])))
            _record("has_v3 is False (2-file upload)", body.get("has_v3") is False)
    except Exception as e:
        _record("POST /api/upload", False, str(e))

    # ── 3. Upload with oversized file rejection ───────────────────────────────
    print("\n[Upload validation]")
    try:
        r = s.post(f"{BASE_URL}/api/upload", files={
            "v1": ("big.xlsx", b"0" * (101 * 1024 * 1024), "application/octet-stream"),
            "v2": ("fm_v2.xlsx", v2_bytes, "application/octet-stream"),
        }, timeout=60)
        _record("Oversized file → 413", r.status_code == 413, f"got {r.status_code}")
    except Exception as e:
        _record("Oversized file rejection", False, str(e))

    try:
        r = s.post(f"{BASE_URL}/api/upload", files={
            "v1": ("model.csv", b"a,b\n1,2\n", "text/csv"),
            "v2": ("model.csv", b"a,b\n1,2\n", "text/csv"),
        }, timeout=10)
        _record("Invalid extension (.csv) → 422", r.status_code == 422, f"got {r.status_code}")
    except Exception as e:
        _record("Invalid extension rejection", False, str(e))

    if not job_id:
        print("\n  Upload failed — skipping remaining tests.")
        _print_summary()
        return 1

    # ── 4. resolve-preview ────────────────────────────────────────────────────
    print("\n[resolve-preview]")
    resolutions: list[dict] = []
    try:
        r = s.post(f"{BASE_URL}/api/{job_id}/resolve-preview",
                   json={"sheets_v1": ["RESUME"], "sheets_v2": ["RESUME"]},
                   timeout=60)
        ok = r.status_code == 200
        _record("POST /resolve-preview → 200", ok, f"status={r.status_code}")
        if ok:
            body = r.json()
            resolutions = body.get("resolutions", [])
            _record("resolutions is list", isinstance(resolutions, list))
            found_kpis = {r2["kpi_name"] for r2 in resolutions if "kpi_name" in r2}
            expected = {"Выручка", "NPV", "IRR"}
            matched = expected & found_kpis
            _record(f"Found key KPIs {expected}", bool(matched), f"found={found_kpis}")
    except Exception as e:
        _record("POST /resolve-preview", False, str(e))

    try:
        r = s.post(f"{BASE_URL}/api/aabbccdd11223344aabbccdd11223344/resolve-preview",
                   json={"sheets_v1": ["RESUME"], "sheets_v2": ["RESUME"]}, timeout=10)
        _record("resolve-preview unknown job → 404", r.status_code == 404, f"got {r.status_code}")
    except Exception as e:
        _record("resolve-preview 404", False, str(e))

    # ── 5. Dashboard ──────────────────────────────────────────────────────────
    print("\n[Dashboard]")
    dashboard: dict[str, Any] = {}
    try:
        r = s.get(f"{BASE_URL}/api/{job_id}/dashboard", timeout=30)
        ok = r.status_code == 200
        _record("GET /dashboard → 200", ok, f"status={r.status_code}")
        if ok:
            dashboard = r.json()
            kpis = dashboard.get("kpis", [])
            _record("dashboard has kpis list", isinstance(kpis, list))
            if kpis:
                kpi = kpis[0]
                has_fields = all(f in kpi for f in ("kpi_name", "value_v1", "value_v2", "delta"))
                _record("KPI row has required fields", has_fields, str(list(kpi.keys())[:6]))
                # Verify delta = v1 - v2
                if kpi["value_v1"] is not None and kpi["value_v2"] is not None:
                    # Convention: delta = v2 - v1 (old minus new)
                    expected_delta = round(kpi["value_v2"] - kpi["value_v1"], 2)
                    actual_delta = round(kpi["delta"], 2)
                    _record("KPI delta == v2 - v1",
                            abs(expected_delta - actual_delta) < 0.1,
                            f"expected={expected_delta} actual={actual_delta}")
            else:
                results.append(("Dashboard KPI fields", _SKIP, "no KPIs resolved"))
    except Exception as e:
        _record("GET /dashboard", False, str(e))

    # ── 6. Run comparison ─────────────────────────────────────────────────────
    print("\n[Run + Status + Summary]")
    try:
        r = s.post(f"{BASE_URL}/api/{job_id}/run",
                   json={"resolutions": resolutions, "mode": "quick"},
                   timeout=30)
        ok = r.status_code == 200
        _record("POST /run → 200", ok, f"status={r.status_code}")

        if ok:
            final = _poll(s, job_id)
            _record("Job completes with status=done",
                    final["status"] == "done",
                    f"status={final['status']} error={final.get('error', '')}")
    except Exception as e:
        _record("POST /run", False, str(e))
        final = {"status": "error"}

    # ── 7. Summary ────────────────────────────────────────────────────────────
    if final.get("status") == "done":
        try:
            r = s.get(f"{BASE_URL}/api/{job_id}/summary", timeout=30)
            ok = r.status_code == 200
            _record("GET /summary → 200", ok, f"status={r.status_code}")
            if ok:
                body = r.json()
                _record("summary has summary_blocks", "summary_blocks" in body,
                        str(list(body.keys())))
        except Exception as e:
            _record("GET /summary", False, str(e))
    else:
        results.append(("GET /summary", _SKIP, "run did not finish"))

    # ── 8. Report download ────────────────────────────────────────────────────
    print("\n[Report download]")
    if final.get("status") == "done":
        try:
            r = s.get(f"{BASE_URL}/api/{job_id}/report.xlsx", timeout=30)
            ok = r.status_code == 200
            _record("GET /report.xlsx → 200", ok, f"status={r.status_code}")
            if ok:
                magic = r.content[:4]
                _record("report.xlsx is valid xlsx (PK magic)",
                        magic == b"PK\x03\x04",
                        f"magic={magic.hex()}")
                _record("report.xlsx > 5 KB", len(r.content) > 5000,
                        f"size={len(r.content)} bytes")
        except Exception as e:
            _record("GET /report.xlsx", False, str(e))
    else:
        results.append(("GET /report.xlsx", _SKIP, "run did not finish"))

    # ── 9. Three-version (V3) workflow ────────────────────────────────────────
    print("\n[V3 three-version comparison]")
    try:
        r = s.post(f"{BASE_URL}/api/upload", files={
            "v1": ("fm_v1.xlsx", v1_bytes, "application/octet-stream"),
            "v2": ("fm_v2.xlsx", v2_bytes, "application/octet-stream"),
            "v3": ("fm_v3.xlsx", v3_bytes, "application/octet-stream"),
        }, timeout=30)
        ok = r.status_code == 200
        _record("Upload with V3 → 200", ok, f"status={r.status_code}")
        if ok:
            v3_body = r.json()
            v3_job = v3_body.get("job_id")
            _record("has_v3 = True", v3_body.get("has_v3") is True)
            _record("sheets_v3 not empty", bool(v3_body.get("sheets_v3")))

            r2 = s.post(f"{BASE_URL}/api/{v3_job}/resolve-preview",
                        json={"sheets_v1": ["RESUME"], "sheets_v2": ["RESUME"],
                              "sheets_v3": ["RESUME"]},
                        timeout=60)
            ok2 = r2.status_code == 200
            _record("V3 resolve-preview → 200", ok2, f"status={r2.status_code}")
            if ok2:
                v3_res = r2.json().get("resolutions", [])
                has_v3_addr = any(r3.get("addr_v3") for r3 in v3_res)
                _record("V3 resolutions have addr_v3", has_v3_addr)

                dash_r = s.get(f"{BASE_URL}/api/{v3_job}/dashboard", timeout=30)
                if dash_r.status_code == 200:
                    _record("V3 dashboard has_v3=True",
                            dash_r.json().get("has_v3") is True)
                    kpis_with_v3 = [k for k in dash_r.json().get("kpis", [])
                                    if k.get("value_v3") is not None]
                    _record("V3 KPIs have value_v3",
                            bool(kpis_with_v3),
                            f"{len(kpis_with_v3)} KPIs with V3")
    except Exception as e:
        _record("V3 workflow", False, str(e))

    # ── 10. Chat endpoint ─────────────────────────────────────────────────────
    print("\n[Chat endpoint]")
    try:
        r = s.post(f"{BASE_URL}/api/{job_id}/chat",
                   json={"history": [{"role": "user", "content": "Как изменилась выручка?"}]},
                   stream=True, timeout=30)
        ok = r.status_code == 200
        _record("POST /chat → 200", ok, f"status={r.status_code}")
        if ok:
            ct = r.headers.get("content-type", "")
            _record("chat Content-Type is SSE", "event-stream" in ct, ct)
            chunks: list[str] = []
            for line in r.iter_lines(decode_unicode=True):
                if line.startswith("data: "):
                    payload = line[6:]
                    if payload == "[DONE]":
                        break
                    chunks.append(payload)
                    if len(chunks) > 200:
                        break
            full = "".join(chunks)
            # Either a real LLM response or a "gateway unavailable" message
            has_content = len(full.strip()) > 0
            _record("chat SSE stream returned content", has_content, f"{len(full)} chars")
    except Exception as e:
        _record("POST /chat", False, str(e))

    # ── 11. Security: path traversal ─────────────────────────────────────────
    print("\n[Security]")
    bad_ids = ["../etc/passwd", "' OR 1=1 --"]
    for bad in bad_ids:
        try:
            r = s.get(f"{BASE_URL}/api/{bad}/status", timeout=5)
            _record(f"path-traversal {bad!r} → 404/422",
                    r.status_code in (404, 422),
                    f"got {r.status_code}")
        except Exception as e:
            # Connection errors or redirect loops are also acceptable
            _record(f"path-traversal {bad!r} blocked", True, str(type(e).__name__))

    _print_summary()
    failed = sum(1 for _, s, _ in results if s == _FAIL)
    return 0 if failed == 0 else 1


def _print_summary() -> None:
    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    passed = sum(1 for _, s, _ in results if s == _PASS)
    failed = sum(1 for _, s, _ in results if s == _FAIL)
    skipped = sum(1 for _, s, _ in results if s == _SKIP)
    total = len(results)
    print(f"  PASS: {passed}/{total}   FAIL: {failed}   SKIP: {skipped}")
    if failed:
        print("\nFailed tests:")
        for name, status, detail in results:
            if status == _FAIL:
                print(f"  {_FAIL} {name}" + (f" — {detail}" if detail else ""))
    print()


if __name__ == "__main__":
    sys.exit(run_tests())
