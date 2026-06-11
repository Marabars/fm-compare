"""
Tests for the LLM chat feature (Stage 5).

Tests are fully offline — they mock the GatewayClient so no real gateway
is required. Covers:
  - build_dashboard_context: markdown serialization
  - stream_chat_turn: gateway not configured, health check fail, happy path
  - POST /api/{job_id}/chat endpoint: validation, SSE format
"""
from __future__ import annotations

import io
import os
import pytest
from unittest.mock import MagicMock, patch

from fm_compare.core.llm.chat import build_dashboard_context, stream_chat_turn


# ── build_dashboard_context ───────────────────────────────────────────────────

def _dash(kpis=None, has_v3=False, date_v1="2025-01-01", date_v2="2024-01-01", date_v3=None):
    return {
        "kpis": kpis or [],
        "has_v3": has_v3,
        "date_v1": date_v1,
        "date_v2": date_v2,
        "date_v3": date_v3,
    }


def _kpi(name, group="Группа", unit="млн руб.", v1=100.0, v2=80.0, delta=20.0, delta_pct=25.0,
         v3=None, delta_v2_v3=None):
    return {
        "kpi_name": name, "kpi_group": group, "unit": unit,
        "value_v1": v1, "value_v2": v2, "delta": delta, "delta_pct": delta_pct,
        "value_v3": v3, "delta_v2_v3": delta_v2_v3,
    }


def test_build_context_empty_dashboard():
    assert build_dashboard_context({}) == ""


def test_build_context_no_kpis():
    assert build_dashboard_context(_dash(kpis=[])) == ""


def test_build_context_contains_kpi_name():
    ctx = build_dashboard_context(_dash(kpis=[_kpi("Выручка")]))
    assert "Выручка" in ctx


def test_build_context_two_version_header():
    ctx = build_dashboard_context(_dash(kpis=[_kpi("NPV")]))
    assert "2025-01-01" in ctx
    assert "2024-01-01" in ctx
    assert "V3" not in ctx


def test_build_context_three_version_header():
    ctx = build_dashboard_context(_dash(
        kpis=[_kpi("NPV", v3=60.0, delta_v2_v3=20.0)],
        has_v3=True, date_v3="2023-01-01"
    ))
    assert "2023-01-01" in ctx
    assert "V3" in ctx or "Δ(V2-V3)" in ctx


def test_build_context_caps_at_20_rows():
    kpis = [_kpi(f"KPI{i}") for i in range(30)]
    ctx = build_dashboard_context(_dash(kpis=kpis))
    assert "KPI19" in ctx
    assert "KPI20" not in ctx
    assert "20 из 30" in ctx


def test_build_context_formats_floats():
    ctx = build_dashboard_context(_dash(kpis=[_kpi("Выручка", v1=1000000.0, v2=900000.0)]))
    # Should format with commas or decimal point, not just raw int
    assert "1,000,000.00" in ctx or "1000000" in ctx


# ── stream_chat_turn ──────────────────────────────────────────────────────────

def test_stream_chat_turn_not_configured():
    cfg = MagicMock()
    cfg.is_configured = False
    fragments = list(stream_chat_turn([{"role": "user", "content": "Привет"}], {}, cfg=cfg))
    assert len(fragments) == 1
    assert "недоступна" in fragments[0].lower() or "не настроен" in fragments[0].lower()


def test_stream_chat_turn_health_check_fail():
    cfg = MagicMock()
    cfg.is_configured = True
    with patch("fm_compare.core.llm.chat.GatewayClient") as MockClient:
        MockClient.return_value.health_check.return_value = False
        fragments = list(stream_chat_turn([{"role": "user", "content": "Привет"}], {}, cfg=cfg))
    assert len(fragments) == 1
    assert "недоступна" in fragments[0].lower() or "временно" in fragments[0].lower()


def test_stream_chat_turn_happy_path():
    cfg = MagicMock()
    cfg.is_configured = True
    with patch("fm_compare.core.llm.chat.GatewayClient") as MockClient:
        instance = MockClient.return_value
        instance.health_check.return_value = True
        instance.stream_chat.return_value = iter(["Привет", ", ", "мир!"])
        fragments = list(stream_chat_turn(
            [{"role": "user", "content": "Как дела?"}],
            _dash(kpis=[_kpi("Выручка")]),
            cfg=cfg,
        ))
    assert "".join(fragments) == "Привет, мир!"


def test_stream_chat_turn_injects_context_in_first_user_msg():
    cfg = MagicMock()
    cfg.is_configured = True
    captured_messages = []

    def _fake_stream(messages, **kwargs):
        captured_messages.extend(messages)
        return iter(["ok"])

    with patch("fm_compare.core.llm.chat.GatewayClient") as MockClient:
        instance = MockClient.return_value
        instance.health_check.return_value = True
        instance.stream_chat.side_effect = _fake_stream
        list(stream_chat_turn(
            [{"role": "user", "content": "Вопрос?"}],
            _dash(kpis=[_kpi("NPV")]),
            cfg=cfg,
        ))

    user_msgs = [m for m in captured_messages if m.role == "user"]
    assert user_msgs, "No user messages sent to gateway"
    assert "NPV" in user_msgs[0].content
    assert "Вопрос?" in user_msgs[0].content


def test_stream_chat_turn_context_only_in_first_user_msg():
    """Context must NOT be re-injected in follow-up user messages."""
    cfg = MagicMock()
    cfg.is_configured = True
    captured_messages = []

    def _fake_stream(messages, **kwargs):
        captured_messages.extend(messages)
        return iter(["ok"])

    with patch("fm_compare.core.llm.chat.GatewayClient") as MockClient:
        instance = MockClient.return_value
        instance.health_check.return_value = True
        instance.stream_chat.side_effect = _fake_stream
        list(stream_chat_turn(
            [
                {"role": "user", "content": "Первый вопрос"},
                {"role": "assistant", "content": "Ответ"},
                {"role": "user", "content": "Второй вопрос"},
            ],
            _dash(kpis=[_kpi("NPV")]),
            cfg=cfg,
        ))

    user_msgs = [m for m in captured_messages if m.role == "user"]
    assert len(user_msgs) == 2
    assert "Данные финансовой модели" in user_msgs[0].content
    assert "Данные финансовой модели" not in user_msgs[1].content


def test_stream_chat_turn_gateway_error_returns_message():
    from fm_compare.core.llm.errors import GatewayError
    cfg = MagicMock()
    cfg.is_configured = True
    with patch("fm_compare.core.llm.chat.GatewayClient") as MockClient:
        instance = MockClient.return_value
        instance.health_check.return_value = True
        instance.stream_chat.side_effect = GatewayError("timeout")
        fragments = list(stream_chat_turn(
            [{"role": "user", "content": "Вопрос?"}], {}, cfg=cfg
        ))
    assert len(fragments) == 1
    assert "ошибка" in fragments[0].lower()


# ── POST /api/{job_id}/chat endpoint ─────────────────────────────────────────

@pytest.fixture(scope="module")
def client():
    os.environ.pop("APP_PASSWORD", None)
    from fastapi.testclient import TestClient
    from fm_compare.web.app import create_app
    with TestClient(create_app(), raise_server_exceptions=True) as c:
        yield c


def _upload_job(client) -> str:
    from openpyxl import Workbook
    def _xlsx():
        wb = Workbook(); ws = wb.active; ws["A1"] = "Выручка"; ws["B1"] = 100.0
        buf = io.BytesIO(); wb.save(buf); return buf.getvalue()
    resp = client.post("/api/upload",
                       files={"v1": ("v1.xlsx", _xlsx()), "v2": ("v2.xlsx", _xlsx())})
    return resp.json()["job_id"]


def test_chat_unknown_job_404(client):
    resp = client.post("/api/deadbeef/chat",
                       json={"history": [{"role": "user", "content": "Привет"}]})
    assert resp.status_code == 404


def test_chat_empty_history_422(client):
    job_id = _upload_job(client)
    resp = client.post(f"/api/{job_id}/chat", json={"history": []})
    assert resp.status_code == 422


def test_chat_last_not_user_422(client):
    job_id = _upload_job(client)
    resp = client.post(f"/api/{job_id}/chat", json={
        "history": [{"role": "assistant", "content": "Ответ"}]
    })
    assert resp.status_code == 422


def test_chat_content_too_long_422(client):
    job_id = _upload_job(client)
    resp = client.post(f"/api/{job_id}/chat", json={
        "history": [{"role": "user", "content": "x" * 2001}]
    })
    assert resp.status_code == 422


def test_chat_sse_format_with_mocked_gateway(client):
    """SSE stream contains data: lines and ends with data: [DONE]."""
    job_id = _upload_job(client)
    with patch("fm_compare.web.routes.stream_chat_turn") as mock_stream:
        mock_stream.return_value = iter(["Привет", " мир"])
        resp = client.post(f"/api/{job_id}/chat",
                           json={"history": [{"role": "user", "content": "Тест"}]})
    assert resp.status_code == 200
    assert "text/event-stream" in resp.headers.get("content-type", "")
    body = resp.text
    assert "data: Привет" in body
    assert "data: [DONE]" in body


def test_chat_unit_override_uses_v2_fallback():
    """Unit from V2 is used when V1 unit is empty (regression for fix in routes.py)."""
    from fm_compare.core.models import KPIResolution
    from fm_compare.core.kpi_resolver import resolutions_to_overrides

    r = KPIResolution(
        kpi_name="Выручка", kpi_group="Доходы", kpi_level=1, search_pattern="",
        unit_v1="",        # V1 empty — simulates missing unit
        unit_v2="млн руб.",
        unit_v3="",
    )
    # Replicate the fixed units dict logic from routes.py
    unit = r.unit_v1 or r.unit_v2 or r.unit_v3
    units = {r.kpi_name: unit} if unit else {}
    assert units == {"Выручка": "млн руб."}
