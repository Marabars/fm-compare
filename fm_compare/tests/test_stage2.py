"""
Stage 2 tests: recalc engine + sensitivity analysis.

recalc.py tests mock LibreOffice (soffice binary absent in CI) and validate
the patch logic using openpyxl in-memory workbooks.

sensitivity.py + API tests mock recalc_with_overrides to return known data.
"""
from __future__ import annotations

import io
import os
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest


# ── helpers ───────────────────────────────────────────────────────────────────

def _make_xlsx(values: dict[tuple[str, int, int], object] | None = None) -> bytes:
    """Build minimal in-memory xlsx with openpyxl.  values: {(sheet, row, col): value}"""
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    sheets: set[str] = set()
    if values:
        for sheet, *_ in values:
            sheets.add(sheet)
    for sh in (sheets or {"Sheet1"}):
        wb.create_sheet(sh)
    if values:
        for (sheet, row, col), val in values.items():
            wb[sheet].cell(row=row, column=col).value = val
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_xlsx(path: Path, values: dict) -> None:
    path.write_bytes(_make_xlsx(values))


# ── _parse_addr_str ────────────────────────────────────────────────────────────

class TestParseAddrStr:

    def test_basic(self):
        from fm_compare.core.recalc import _parse_addr_str
        sheet, row, col = _parse_addr_str("PRICE!B5")
        assert sheet == "PRICE"
        assert row == 5
        assert col == 2   # B = column 2

    def test_multi_letter_col(self):
        from fm_compare.core.recalc import _parse_addr_str
        _, _, col = _parse_addr_str("Sheet1!AA1")
        assert col == 27  # AA = 27

    def test_no_exclamation_raises(self):
        from fm_compare.core.recalc import _parse_addr_str, RecalcError
        with pytest.raises(RecalcError):
            _parse_addr_str("A1")

    def test_bad_cell_raises(self):
        from fm_compare.core.recalc import _parse_addr_str, RecalcError
        with pytest.raises(RecalcError):
            _parse_addr_str("Sheet!notacell")


# ── addr_to_str round-trip ────────────────────────────────────────────────────

def test_addr_to_str_round_trip():
    from fm_compare.core.models import CellAddress
    from fm_compare.core.recalc import addr_to_str, _parse_addr_str
    addr = CellAddress(sheet="PRICE", row=5, col=2)
    s = addr_to_str(addr)
    assert s == "PRICE!B5"
    sheet, row, col = _parse_addr_str(s)
    assert (sheet, row, col) == ("PRICE", 5, 2)


# ── _patch_overrides ───────────────────────────────────────────────────────────

class TestPatchOverrides:

    def test_patches_string_key(self, tmp_path):
        from fm_compare.core.recalc import _patch_overrides
        import openpyxl

        p = tmp_path / "test.xlsx"
        _write_xlsx(p, {("PRICE", 5, 2): 100.0})

        _patch_overrides(p, {"PRICE!B5": 200.0})

        wb = openpyxl.load_workbook(str(p), data_only=True)
        assert wb["PRICE"].cell(5, 2).value == 200.0

    def test_patches_cell_address_via_addr_to_str(self, tmp_path):
        """addr_to_str converts CellAddress to a string key usable by _patch_overrides."""
        from fm_compare.core.models import CellAddress
        from fm_compare.core.recalc import _patch_overrides, addr_to_str
        import openpyxl

        p = tmp_path / "test.xlsx"
        _write_xlsx(p, {("Sheet1", 1, 1): 42.0})

        addr = CellAddress(sheet="Sheet1", row=1, col=1)
        _patch_overrides(p, {addr_to_str(addr): 99.0})

        wb = openpyxl.load_workbook(str(p), data_only=True)
        assert wb["Sheet1"].cell(1, 1).value == 99.0

    def test_ignores_unknown_sheet(self, tmp_path):
        from fm_compare.core.recalc import _patch_overrides
        import openpyxl

        p = tmp_path / "test.xlsx"
        _write_xlsx(p, {("Sheet1", 1, 1): 1.0})

        # Should not raise, just skip
        _patch_overrides(p, {"MISSING!A1": 999.0})
        wb = openpyxl.load_workbook(str(p), data_only=True)
        assert wb["Sheet1"].cell(1, 1).value == 1.0


# ── recalc_with_overrides (mocked LibreOffice) ─────────────────────────────────

class TestRecalcWithOverrides:

    def test_raises_when_soffice_missing(self, tmp_path):
        from fm_compare.core.recalc import recalc_with_overrides, RecalcError
        p = tmp_path / "fm.xlsx"
        _write_xlsx(p, {("Sheet1", 1, 1): 1.0})

        with patch("fm_compare.core.recalc.shutil.which", return_value=None):
            with pytest.raises(RecalcError, match="LibreOffice not found"):
                recalc_with_overrides(p, {})

    def test_full_flow_mocked(self, tmp_path):
        """End-to-end with mocked soffice: patches, calls LO, reads result."""
        from fm_compare.core.recalc import recalc_with_overrides
        from fm_compare.core.excel_reader import WorkbookData, SheetData

        src = tmp_path / "fm.xlsx"
        _write_xlsx(src, {("Sheet1", 1, 1): 50.0})

        # Build a fake "recalculated" xlsx that _run_libreoffice would emit
        recalc_xlsx = _make_xlsx({("Sheet1", 1, 1): 200.0})

        def fake_run_lo(soffice, src_path, out_dir, timeout):
            (out_dir / src_path.name).write_bytes(recalc_xlsx)

        with patch("fm_compare.core.recalc.shutil.which", return_value="/usr/bin/soffice"):
            with patch("fm_compare.core.recalc._run_libreoffice", side_effect=fake_run_lo):
                wb = recalc_with_overrides(src, {"Sheet1!A1": 99.0})

        assert "Sheet1" in wb.sheets
        cell = wb.sheets["Sheet1"].cells.get((1, 1))
        assert cell is not None
        assert cell.value == 200.0   # the fake recalculated value


# ── sensitivity run_scenarios (mocked recalc) ────────────────────────────────

class TestRunScenarios:

    def _wb_with_kpi(self, value: float) -> object:
        """Return a WorkbookData with 'DB'!C10 = value."""
        from fm_compare.core.excel_reader import WorkbookData, SheetData, CellData
        from fm_compare.core.models import CellAddress
        cell = CellData(
            address=CellAddress(sheet="DB", row=10, col=3),
            value=value, formula=None, data_type="n",
            comment=None, is_hidden_row=False,
        )
        sheet = SheetData(name="DB", cells={(10, 3): cell}, max_row=10, max_col=3)
        wb = WorkbookData()
        wb.sheets["DB"] = sheet
        return wb

    def test_returns_base_and_variations(self, tmp_path):
        from fm_compare.core.models import CellAddress, SensitivityInput
        from fm_compare.core.sensitivity import run_scenarios

        path = tmp_path / "fm.xlsx"
        _write_xlsx(path, {("PRICE", 5, 2): 150_000.0})

        inp = SensitivityInput(
            name="Цена",
            addr=CellAddress(sheet="PRICE", row=5, col=2),
            base_value=150_000.0,
            values=[120_000.0, 150_000.0, 180_000.0],
        )
        kpi_addrs = {"NPV": CellAddress(sheet="DB", row=10, col=3)}

        call_count = [0]

        def fake_recalc(p, overrides, timeout=90):
            kpi_val = 100.0 + call_count[0] * 20.0
            call_count[0] += 1
            return self._wb_with_kpi(kpi_val)

        with patch("fm_compare.core.sensitivity.recalc_with_overrides", side_effect=fake_recalc):
            result = run_scenarios(path, [inp], kpi_addrs)

        assert result.input_names == ["Цена"]
        assert result.kpi_names == ["NPV"]
        assert result.base_scenario.label == "base"
        # base + 2 non-base variations (150_000 reuses base)
        assert len(result.scenarios) == 3

    def test_empty_inputs_raises(self, tmp_path):
        from fm_compare.core.models import CellAddress
        from fm_compare.core.sensitivity import run_scenarios

        path = tmp_path / "fm.xlsx"
        _write_xlsx(path, {("Sheet1", 1, 1): 1.0})
        with pytest.raises(ValueError, match="SensitivityInput"):
            run_scenarios(path, [], {"NPV": CellAddress("DB", 10, 3)})

    def test_empty_kpis_raises(self, tmp_path):
        from fm_compare.core.models import CellAddress, SensitivityInput
        from fm_compare.core.sensitivity import run_scenarios

        path = tmp_path / "fm.xlsx"
        _write_xlsx(path, {("Sheet1", 1, 1): 1.0})
        inp = SensitivityInput("X", CellAddress("Sheet1", 1, 1), base_value=1.0, values=[1.0])
        with pytest.raises(ValueError, match="KPI"):
            run_scenarios(path, [inp], {})


# ── sensitivity API endpoints ─────────────────────────────────────────────────

@pytest.fixture(scope="module")
def client():
    os.environ.pop("APP_PASSWORD", None)
    from fastapi.testclient import TestClient
    from fm_compare.web.app import create_app
    with TestClient(create_app(), raise_server_exceptions=True) as c:
        yield c


def _upload_job(client) -> str:
    data = _make_xlsx({("PRICE", 5, 2): 150_000.0, ("DB", 10, 3): 1_000_000.0})
    resp = client.post(
        "/api/upload",
        files={"v1": ("fm.xlsx", data, "application/octet-stream"),
               "v2": ("fm.xlsx", data, "application/octet-stream")},
    )
    assert resp.status_code == 200
    return resp.json()["job_id"]


class TestSensitivityEndpoints:

    def test_idle_before_run(self, client):
        job_id = _upload_job(client)
        resp = client.get(f"/api/{job_id}/sensitivity")
        assert resp.status_code == 200
        assert resp.json()["status"] == "idle"

    def test_post_starts_run(self, client):
        import fm_compare.core.sensitivity  # ensure module imported before patch target resolves
        job_id = _upload_job(client)
        body = {
            "inputs": [{"name": "Цена", "addr": "PRICE!B5",
                        "unit": "руб./кв.м", "base_value": 150000,
                        "values": [120000, 150000, 180000]}],
            "kpi_addrs": [{"name": "NPV", "addr": "DB!C10"}],
            "timeout": 5,
        }
        # Patch recalc so the sensitivity finishes without LibreOffice
        from fm_compare.core.excel_reader import WorkbookData, SheetData, CellData
        from fm_compare.core.models import CellAddress

        def fake_recalc(p, overrides, timeout=90):
            cell = CellData(CellAddress("DB", 10, 3), 1e9, None, "n", None, False)
            sh = SheetData(name="DB", cells={(10, 3): cell}, max_row=10, max_col=3)
            wb = WorkbookData()
            wb.sheets["DB"] = sh
            return wb

        with patch("fm_compare.core.sensitivity.recalc_with_overrides", side_effect=fake_recalc):
            resp = client.post(f"/api/{job_id}/sensitivity",
                               json=body,
                               headers={"Content-Type": "application/json"})
        assert resp.status_code == 200
        assert resp.json()["status"] == "running"

    def test_post_unknown_job_404(self, client):
        resp = client.post("/api/nonexistent/sensitivity", json={
            "inputs": [{"name": "x", "addr": "S!A1", "base_value": 1, "values": [1]}],
            "kpi_addrs": [{"name": "NPV", "addr": "S!B1"}],
        })
        assert resp.status_code == 404

    def test_post_no_inputs_422(self, client):
        job_id = _upload_job(client)
        resp = client.post(f"/api/{job_id}/sensitivity", json={
            "inputs": [],
            "kpi_addrs": [{"name": "NPV", "addr": "S!B1"}],
        })
        assert resp.status_code == 422

    def test_post_no_kpis_422(self, client):
        job_id = _upload_job(client)
        resp = client.post(f"/api/{job_id}/sensitivity", json={
            "inputs": [{"name": "x", "addr": "S!A1", "base_value": 1, "values": [1, 2]}],
            "kpi_addrs": [],
        })
        assert resp.status_code == 422

    def test_post_bad_addr_422(self, client):
        job_id = _upload_job(client)
        resp = client.post(f"/api/{job_id}/sensitivity", json={
            "inputs": [{"name": "x", "addr": "not-an-addr", "base_value": 1, "values": [1]}],
            "kpi_addrs": [{"name": "NPV", "addr": "S!B1"}],
        })
        assert resp.status_code == 422
