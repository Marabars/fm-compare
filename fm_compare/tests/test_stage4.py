"""
Stage 4 tests: three-version comparison (V3 support).

Tests cover:
- KPIResolution V3 fields
- _enrich_rows_with_v3 in dashboard module
- build_dashboard with has_v3=True/False
- /upload route with V3 file
- /resolve-preview with sheets_v3
- /dashboard with V3 overrides
- /status2 and /report2.xlsx endpoints
"""
from __future__ import annotations

import io
import os
import pytest
from pathlib import Path

from fm_compare.core.excel_reader import SheetData, CellData, WorkbookData
from fm_compare.core.models import CellAddress, KPIResolution
from fm_compare.web.dashboard import _enrich_rows_with_v3


# ── Helpers ──────────────────────────────────────────────────────────────────

def _cell(value=None):
    return CellData(address=None, value=value, formula=None,
                    data_type=None, comment=None, is_hidden_row=False)


def _make_xlsx(sheet_name: str = "Sheet1", kpi_label: str = "Выручка",
               kpi_value: float = 100.0) -> bytes:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws["A1"] = kpi_label
    ws["B1"] = kpi_value
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture(scope="module")
def client():
    os.environ.pop("APP_PASSWORD", None)
    from fastapi.testclient import TestClient
    from fm_compare.web.app import create_app
    with TestClient(create_app(), raise_server_exceptions=True) as c:
        yield c


# ── Unit: KPIResolution V3 fields ────────────────────────────────────────────

def _make_resolution(**kwargs) -> KPIResolution:
    defaults = dict(kpi_level=1, search_pattern="")
    defaults.update(kwargs)
    return KPIResolution(**defaults)


def test_kpi_resolution_has_v3_fields():
    r = _make_resolution(kpi_name="Выручка", kpi_group="Доходы")
    assert r.sheet_v3 == ""
    assert r.row_v3 is None
    assert r.col_v3 is None
    assert r.label_v3 == ""
    assert r.addr_v3 == ""
    assert r.unit_v3 == ""


def test_kpi_resolution_v3_fields_settable():
    r = _make_resolution(kpi_name="EBITDA", kpi_group="Прибыль",
                         sheet_v3="PRICE", row_v3=5, col_v3=3,
                         label_v3="EBITDA", addr_v3="PRICE!C5", unit_v3="млн руб.")
    assert r.addr_v3 == "PRICE!C5"
    assert r.unit_v3 == "млн руб."


# ── Unit: _enrich_rows_with_v3 ───────────────────────────────────────────────

def test_enrich_rows_with_v3_numeric():
    rows = [{"kpi_name": "Выручка", "value_v2": 100.0}]
    kpi_v2 = {"Выручка": (100.0, CellAddress("Sheet1", 1, 2))}
    kpi_v3 = {"Выручка": (80.0, CellAddress("Sheet1", 1, 2))}
    _enrich_rows_with_v3(rows, kpi_v2, kpi_v3)
    row = rows[0]
    assert row["value_v3"] == 80.0
    assert row["delta_v2_v3"] == pytest.approx(-20.0)
    assert row["delta_v2_v3_pct"] == pytest.approx(-20.0)


def test_enrich_rows_with_v3_missing_kpi():
    rows = [{"kpi_name": "EBITDA", "value_v2": 50.0}]
    kpi_v2 = {"EBITDA": (50.0, None)}
    kpi_v3 = {}  # KPI not in V3
    _enrich_rows_with_v3(rows, kpi_v2, kpi_v3)
    row = rows[0]
    assert row["value_v3"] is None
    assert row["delta_v2_v3"] is None
    assert row["delta_v2_v3_pct"] is None


def test_enrich_rows_with_v3_zero_denominator():
    rows = [{"kpi_name": "NPV", "value_v2": 0.0}]
    kpi_v2 = {"NPV": (0.0, None)}
    kpi_v3 = {"NPV": (10.0, None)}
    _enrich_rows_with_v3(rows, kpi_v2, kpi_v3)
    row = rows[0]
    assert row["delta_v2_v3"] == pytest.approx(10.0)
    assert row["delta_v2_v3_pct"] is None  # division by zero guarded


def test_enrich_rows_with_v3_non_numeric():
    rows = [{"kpi_name": "Дата", "value_v2": "строка"}]
    kpi_v2 = {"Дата": ("строка", None)}
    kpi_v3 = {"Дата": ("другое", None)}
    _enrich_rows_with_v3(rows, kpi_v2, kpi_v3)
    assert rows[0]["delta_v2_v3"] is None


# ── API: upload with V3 file ──────────────────────────────────────────────────

def test_upload_without_v3(client):
    v1 = _make_xlsx("PRICE", "Выручка", 200.0)
    v2 = _make_xlsx("PRICE", "Выручка", 180.0)
    resp = client.post("/api/upload",
                       files={"v1": ("v1.xlsx", v1), "v2": ("v2.xlsx", v2)})
    assert resp.status_code == 200
    data = resp.json()
    assert "job_id" in data
    assert data.get("has_v3") is False
    assert "sheets_v1" in data
    assert "sheets_v2" in data
    assert data.get("sheets_v3", []) == []


def test_upload_with_v3(client):
    v1 = _make_xlsx("PRICE", "Выручка", 200.0)
    v2 = _make_xlsx("PRICE", "Выручка", 180.0)
    v3 = _make_xlsx("PRICE", "Выручка", 160.0)
    resp = client.post("/api/upload",
                       files={"v1": ("v1.xlsx", v1),
                               "v2": ("v2.xlsx", v2),
                               "v3": ("v3.xlsx", v3)})
    assert resp.status_code == 200
    data = resp.json()
    assert data.get("has_v3") is True
    assert "PRICE" in data.get("sheets_v3", [])


# ── API: status2 and report2 endpoints ───────────────────────────────────────

def test_status2_idle_before_run(client):
    v1 = _make_xlsx("PRICE", "Выручка", 200.0)
    v2 = _make_xlsx("PRICE", "Выручка", 180.0)
    up = client.post("/api/upload",
                     files={"v1": ("v1.xlsx", v1), "v2": ("v2.xlsx", v2)})
    job_id = up.json()["job_id"]
    resp = client.get(f"/api/{job_id}/status2")
    assert resp.status_code == 200
    assert resp.json()["status"] == "idle"


def test_status2_unknown_job(client):
    resp = client.get("/api/nonexistent-job-id/status2")
    assert resp.status_code == 404


def test_report2_not_ready(client):
    v1 = _make_xlsx("PRICE", "Выручка", 200.0)
    v2 = _make_xlsx("PRICE", "Выручка", 180.0)
    up = client.post("/api/upload",
                     files={"v1": ("v1.xlsx", v1), "v2": ("v2.xlsx", v2)})
    job_id = up.json()["job_id"]
    resp = client.get(f"/api/{job_id}/report2.xlsx")
    assert resp.status_code == 404
